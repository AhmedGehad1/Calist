"""Tests for the pure logic in calist.py. Run with: python -m pytest

These cover the parts with no I/O: filename parsing, value cleaning, ordering,
de-duplication, second-row generation and pre-flight classification, plus a
couple of end-to-end runs against workbooks built on the fly.
"""

import logging
import os
import re
import threading
import zipfile
from pathlib import Path
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

import calist
from device_config import DEVICE_CONFIGS, form


# ── extract_device_code ───────────────────────────────────────────────────────

def test_code_after_first_hyphen():
    assert calist.extract_device_code("Clinic-AGH001.xlsx") == "AGH"


def test_code_without_hyphen_uses_whole_stem():
    assert calist.extract_device_code("VNT023.xlsx") == "VNT"


def test_code_is_uppercased():
    assert calist.extract_device_code("site-agh001.xlsx") == "AGH"


def test_code_is_none_when_no_letters():
    assert calist.extract_device_code("123.xlsx") is None


# ── clean ─────────────────────────────────────────────────────────────────────

def test_clean_none_is_blank():
    assert calist.clean(None) == ""


def test_clean_strips_text():
    assert calist.clean("  ABC  ") == "ABC"


def test_clean_stringifies_numbers_as_is():
    """Deliberate: a numeric cell keeps Excel's float form."""
    assert calist.clean(123456.0) == "123456.0"
    assert calist.clean(1.5) == "1.5"


def test_clean_stringifies_dates_as_is():
    """Deliberate: a real date cell keeps datetime's default repr."""
    assert calist.clean(datetime(2024, 1, 15)) == "2024-01-15 00:00:00"


# ── natural_key ───────────────────────────────────────────────────────────────

def test_numbers_sort_numerically():
    codes = ["AGH10", "AGH9", "AGH1"]
    assert sorted(codes, key=calist.natural_key) == ["AGH1", "AGH9", "AGH10"]


def test_natural_key_never_compares_text_with_numbers():
    # Mixed shapes must not raise TypeError.
    sorted(["AGH1", "AGHX", "1AGH", "AGH"], key=calist.natural_key)


# ── second row ────────────────────────────────────────────────────────────────

def test_second_row_takes_status2_and_rewrites_code():
    parent = {"Code": "AGH001", "Status": "OK", "Status2": "Faulty", "_row_order": 0}
    child = calist.build_second_row(
        parent, {"device_name": "NIBP", "code_replace": ("AGH", "AGCB")})
    assert child["Device"] == "NIBP"
    assert child["Status"] == "Faulty"
    assert child["Code"] == "AGCB001"
    assert child["_row_order"] == 1


def test_second_row_replaces_code_token_only_once():
    parent = {"Code": "AGH-AGH001", "Status": "", "Status2": "", "_row_order": 0}
    child = calist.build_second_row(
        parent, {"device_name": "NIBP", "code_replace": ("AGH", "AGCB")})
    assert child["Code"] == "AGCB-AGH001"


def test_second_row_does_not_mutate_parent():
    parent = {"Code": "AGH001", "Status": "OK", "Status2": "Faulty", "_row_order": 0}
    calist.build_second_row(
        parent, {"device_name": "NIBP", "code_replace": ("AGH", "AGCB")})
    assert parent["Status"] == "OK" and parent["Code"] == "AGH001"


# ── sorting ───────────────────────────────────────────────────────────────────

def test_sub_module_row_follows_its_parent():
    records = [
        {"_group": "AGH002", "_row_order": 0, "Device": "Patient Monitor"},
        {"_group": "AGH001", "_row_order": 1, "Device": "NIBP"},
        {"_group": "AGH001", "_row_order": 0, "Device": "Patient Monitor"},
    ]
    ordered = calist.sort_records(records)
    assert [(r["_group"], r["_row_order"]) for r in ordered] == [
        ("AGH001", 0), ("AGH001", 1), ("AGH002", 0),
    ]


# ── deduplicate_records ───────────────────────────────────────────────────────

def _rec(device, serial):
    return {"Device": device, "S.N": serial}


def test_blank_serials_are_always_kept():
    records = [_rec("ECG", ""), _rec("ECG", ""), _rec("ECG", "")]
    assert len(calist.deduplicate_records(records)) == 3


def test_plain_duplicate_is_dropped():
    records = [_rec("ECG", "SN1"), _rec("ECG", "SN1")]
    assert len(calist.deduplicate_records(records)) == 1


def test_monitor_and_nibp_may_share_a_serial():
    records = [_rec("Patient Monitor", "SN1"), _rec("NIBP", "SN1")]
    assert len(calist.deduplicate_records(records)) == 2


def test_third_record_on_a_shared_serial_is_dropped():
    records = [_rec("Patient Monitor", "SN1"), _rec("NIBP", "SN1"), _rec("ECG", "SN1")]
    kept = calist.deduplicate_records(records)
    assert [r["Device"] for r in kept] == ["Patient Monitor", "NIBP"]


def test_unpaired_devices_may_not_share_a_serial():
    records = [_rec("ECG", "SN1"), _rec("Ultrasound", "SN1")]
    assert len(calist.deduplicate_records(records)) == 1


def test_vital_sign_modules_may_share_a_serial():
    records = [_rec("Vital Sign (SPO2 Module)", "SN1"),
               _rec("Vital Sign (NIBP Module)", "SN1")]
    assert len(calist.deduplicate_records(records)) == 2


# The duplicate warning has to name the two files, because that is what the
# user opens to resolve it — the device type does not identify which form to
# look at when a round holds a dozen of the same model.

def _sourced(device, serial, source):
    return {"Device": device, "S.N": serial, "_source": source}


def test_duplicate_warning_names_both_files(caplog):
    records = [_sourced("ECG", "SN1", "D23-AGH001-0225.xlsx"),
               _sourced("ECG", "SN1", "D23-AGH007-0225.xlsx")]
    with caplog.at_level(logging.WARNING, logger="aggregator"):
        calist.deduplicate_records(records)

    message = caplog.text
    assert "D23-AGH001-0225.xlsx" in message
    assert "D23-AGH007-0225.xlsx" in message
    assert "SN1" in message


def test_duplicate_warning_falls_back_to_the_code_without_a_source():
    # Records built by older callers carry no _source; Code still names the file.
    assert calist.source_name({"Code": "D23-AGH001-0225"}) == "D23-AGH001-0225"
    assert calist.source_name({}) == "?"


def test_a_sub_module_row_reports_the_file_it_came_from():
    """build_second_row rewrites Code, so only _source still names the file."""
    parent = {"Device": "Patient Monitor", "S.N": "SN1", "Code": "Ward-AGH001",
              "_source": "Ward-AGH001.xlsx", "Status2": "Working"}
    extra = calist.build_second_row(
        parent, {"device_name": "NIBP", "code_replace": ("AGH", "NIB")})

    assert extra["Code"] != parent["Code"]          # the token was rewritten
    assert calist.source_name(extra) == "Ward-AGH001.xlsx"


# ── config integrity ──────────────────────────────────────────────────────────

def test_form_produces_the_documented_layout():
    assert form(18, "H32") == {
        "Manufacturer": "E20", "Model": "E18", "S.N": "K18",
        "Location": "K20", "Date": "E16", "Status": "H32",
    }


def test_every_config_has_the_required_fields():
    required = {"Manufacturer", "Model", "S.N", "Location", "Date", "Status"}
    for code, config in DEVICE_CONFIGS.items():
        assert "device_name" in config, code
        assert required <= set(config["cells"]), code


def test_second_row_names_are_covered_by_the_dedup_exemptions():
    """A device rename must not silently break the shared-serial rule."""
    for code, config in DEVICE_CONFIGS.items():
        if "second_row" not in config:
            continue
        pair = frozenset({config["device_name"],
                          config["second_row"]["device_name"]})
        assert pair in calist.ALLOWED_SHARED_SN_PAIRS, code


def test_configs_with_second_row_define_status2():
    for code, config in DEVICE_CONFIGS.items():
        if "second_row" in config:
            assert "Status2" in config["cells"], code


# ── classify_file (pre-flight, no workbook I/O) ───────────────────────────────

def test_classify_recognises_a_known_device():
    outcome = calist.classify_file("Clinic-AC003.xlsx")
    assert outcome.status == calist.READY
    assert outcome.device_code == "AC"
    assert outcome.device_name == "Defibrillator"
    assert outcome.rows == 1
    assert not outcome.is_problem


def test_classify_reports_two_rows_for_module_devices():
    outcome = calist.classify_file("Clinic-AGH001.xlsx")
    assert outcome.rows == 2
    assert "NIBP" in outcome.device_name


def test_classify_flags_an_unknown_code():
    outcome = calist.classify_file("Clinic-ZZZ999.xlsx")
    assert outcome.status == calist.UNKNOWN_CODE
    assert outcome.is_problem
    assert "ZZZ" in outcome.detail


def test_classify_flags_an_unsupported_extension():
    outcome = calist.classify_file("notes.txt")
    assert outcome.status == calist.UNSUPPORTED
    assert outcome.is_problem


def test_classify_does_not_open_the_file():
    """It must work on a path that does not exist — that is the whole point."""
    outcome = calist.classify_file(r"X:\nowhere\Clinic-AF001.xlsx")
    assert outcome.status == calist.READY
    assert outcome.device_name == "ECG"


# ── filename format check ─────────────────────────────────────────────────────

@pytest.mark.parametrize("stem", [
    "G302-AGH001-0425",       # the reference example
    "G302-AGH001-0426",       # a 2026 date
    "A1-BB007-1225",          # shortest site code, December
    "GH1234-VAH012-0125",     # longer codes, January
    "g302-agh001-0425",       # case does not matter
])
def test_accepts_well_formed_names(stem):
    assert calist.check_filename_format(stem) is None


@pytest.mark.parametrize("stem, expect", [
    ("G302-AGH001", "3 parts"),           # too few parts
    ("G302-AGH001-0425-X", "3 parts"),    # too many
    ("302-AGH001-0425", "site code"),     # site must start with a letter
    ("G-AGH001-0425", "site code"),       # site must end with digits
    ("G302-AGH-0425", "device code"),     # device must end with digits
    ("G302-001-0425", "device code"),     # device must start with letters
    ("G302-AGH001-425", "4 digits"),      # date too short
    ("G302-AGH001-04255", "4 digits"),    # date too long
    ("G302-AGH001-ab25", "4 digits"),     # date not numeric
    ("G302-AGH001-1325", "month"),        # month 13
    ("G302-AGH001-0025", "month"),        # month 00
])
def test_rejects_malformed_names_and_says_why(stem, expect):
    problem = calist.check_filename_format(stem)
    assert problem is not None, stem
    assert expect in problem, f"{stem!r} -> {problem!r}"


def test_strict_mode_is_off_by_default():
    """An old-style name stays acceptable unless the caller asks otherwise."""
    assert calist.classify_file("Clinic-AGH001.xlsx").status == calist.READY


def test_strict_mode_flags_a_bad_name():
    outcome = calist.classify_file("Clinic-AGH001.xlsx", strict_names=True)
    assert outcome.status == calist.BAD_FORMAT
    assert outcome.is_problem


def test_strict_mode_passes_a_good_name_through_to_the_device_lookup():
    outcome = calist.classify_file("G302-AGH001-0425.xlsx", strict_names=True)
    assert outcome.status == calist.READY
    assert outcome.device_code == "AGH"
    assert outcome.rows == 2


def test_format_is_checked_before_the_device_code():
    """With strict on, a malformed name reports the format, not the code."""
    outcome = calist.classify_file("nonsense-ZZZ999.xlsx", strict_names=True)
    assert outcome.status == calist.BAD_FORMAT


def test_house_format_still_yields_the_right_device_code():
    assert calist.extract_device_code("G302-AGH001-0425.xlsx") == "AGH"
    assert calist.extract_device_code("G302-VAH012-1226.xlsx") == "VAH"


# ── merged cells ──────────────────────────────────────────────────────────────
#
# The forms draw each answer as a box spanning two columns. A merged range
# stores its value only in the top-left cell, so a map naming the second column
# read blank — which is how the Ultrasound serial went missing when that form
# was re-laid-out onto the E/K columns.

def _merged_form(path, cells, merges):
    wb = Workbook()
    ws = wb.active
    for ref, value in cells.items():
        ws[ref] = value
    for rng in merges:
        ws.merge_cells(rng)
    wb.save(path)
    return str(path)


def test_a_merged_cell_reads_through_to_its_anchor(tmp_path):
    path = _merged_form(tmp_path / "f.xlsx", {"K17": "6061439WX0"}, ["K17:L17"])
    assert calist.read_record(path, {"S.N": "L17"}) == {"S.N": "6061439WX0"}


def test_naming_the_anchor_itself_still_works(tmp_path):
    path = _merged_form(tmp_path / "f.xlsx", {"K17": "6061439WX0"}, ["K17:L17"])
    assert calist.read_record(path, {"S.N": "K17"}) == {"S.N": "6061439WX0"}


def test_an_ordinary_empty_cell_stays_empty(tmp_path):
    """Only cells actually inside a merge resolve; blank means blank."""
    path = _merged_form(tmp_path / "f.xlsx", {"A1": "header"}, ["A1:Z1"])
    assert calist.read_record(path, {"Status": "H30"}) == {"Status": ""}


def test_the_ultrasound_probe_serial_survives_a_merged_layout(tmp_path):
    """The whole BB form, drawn as two-column boxes on the E/K columns."""
    path = _merged_form(
        tmp_path / "G302-BB001-0526.xlsx",
        {"E15": "10-05-2026", "E17": "Versana Essential", "E19": "GE",
         "K17": "6061439WX0", "K19": "Clinics", "K21": "982693WX4",
         "H30": "Working"},
        [f"{c}{r}:{chr(ord(c) + 1)}{r}"
         for r in (15, 17, 19, 21) for c in ("E", "K")])

    record = calist.read_record(path, DEVICE_CONFIGS["BB"]["cells"])

    assert record["S.N"] == "6061439WX0"
    assert record["S.N2"] == "982693WX4"
    assert record["Model"] == "Versana Essential"
    assert record["Manufacturer"] == "GE"
    assert record["Location"] == "Clinics"


def test_the_two_ultrasound_serials_end_up_on_two_lines(tmp_path):
    """The register cell reads '<device serial>\\n(<probe serial>)'."""
    src = tmp_path / "src"
    src.mkdir()
    template = tmp_path / "Device List.xlsx"
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(["No."] + calist.FIELDS, start=1):
        ws.cell(row=3, column=col, value=name)
    wb.save(template)

    form_path = _merged_form(
        src / "G302-BB001-0526.xlsx",
        {"E15": "10-05-2026", "E17": "Versana Essential", "E19": "GE",
         "K17": "6061439WX0", "K19": "Clinics", "K21": "982693WX4",
         "H30": "Working"},
        [f"{c}{r}:{chr(ord(c) + 1)}{r}"
         for r in (15, 17, 19, 21) for c in ("E", "K")])

    result = calist.process_files([form_path], str(template))

    assert result.succeeded
    written = load_workbook(result.output_path).active
    serial_col = calist.TEMPLATE_START_COL + calist.FIELDS.index("S.N")
    assert written.cell(row=calist.TEMPLATE_START_ROW,
                        column=serial_col).value == "6061439WX0\n(982693WX4)"


# ── which sheet gets read ─────────────────────────────────────────────────────
#
# "Always the first tab" was right for every form but one: the X-ray workbook
# opens on an empty "Waveform Dialog" stub left by its macros, so every mapped
# cell read blank on every X-ray.

def _multi_sheet_form(path, sheets):
    """sheets: list of (title, {ref: value}) in file order."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, cells in sheets:
        ws = wb.create_sheet(title)
        for ref, value in cells.items():
            ws[ref] = value
    wb.save(path)
    return str(path)


def test_an_empty_leading_sheet_is_skipped(tmp_path):
    path = _multi_sheet_form(tmp_path / "f.xlsx", [
        ("Waveform Dialog", {}),
        ("Data entry", {"K18": "XR-77341"}),
    ])
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": "XR-77341"}


def test_a_populated_first_sheet_still_wins(tmp_path):
    """Never skip past a sheet that holds data, even if a later one does too."""
    path = _multi_sheet_form(tmp_path / "f.xlsx", [
        ("Device data", {"K18": "the right one"}),
        ("Data entry", {"K18": "the wrong one"}),
    ])
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": "the right one"}


def test_a_workbook_of_empty_sheets_does_not_crash(tmp_path):
    path = _multi_sheet_form(tmp_path / "f.xlsx", [("one", {}), ("two", {})])
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": ""}


def test_the_xray_reads_through_a_stub_tab_to_its_form(tmp_path):
    """The real BF shape: empty macro tab first, merged boxes on Data entry."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Waveform Dialog")
    ws = wb.create_sheet("Data entry")
    for ref, value in {"E16": "10-05-2026", "E18": "Multix Impact",
                       "E20": "Siemens", "K18": "XR-77341",
                       "K20": "TUBE-99812", "K22": "Radiology",
                       "J27": "Pass"}.items():
        ws[ref] = value
    for row in (16, 18, 20, 22):
        ws.merge_cells(f"E{row}:F{row}")
        ws.merge_cells(f"K{row}:L{row}")
    path = tmp_path / "G302-BF001-0526.xlsx"
    wb.save(path)

    record = calist.read_record(str(path), DEVICE_CONFIGS["BF"]["cells"])

    assert record["S.N"] == "XR-77341"
    assert record["S.N2"] == "TUBE-99812"
    assert record["Model"] == "Multix Impact"
    assert record["Manufacturer"] == "Siemens"
    assert record["Location"] == "Radiology"
    assert record["Status"] == "Pass"


# ── reading the package directly ──────────────────────────────────────────────
#
# The reader goes at the sheet XML rather than through load_workbook, which
# took ~500ms a form because it parses every tab and the whole styles table to
# reach seven cells. These pin the parts of that which a fixture built with
# openpyxl cannot reach on its own.

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"


def _handmade_xlsx(path, sheets, shared=None, styles=None):
    """Write an .xlsx entry by entry, without openpyxl.

    Two shapes matter here and openpyxl cannot produce either. It writes every
    string as an *inline* string, so a fixture built with it never exercises
    the shared-string table — which is what Excel itself writes and what every
    real form uses. And it cannot emit the ``<sheet r:id="">`` entries an older
    macro-enabled workbook carries for its VBA modules.

    ``sheets`` is a list of ``(name, rel_id, kind, body_xml)``. A blank rel_id
    writes the broken entry, with no part behind it.
    """
    entries, rels = [], []
    for index, (name, rel_id, kind, body) in enumerate(sheets, start=1):
        if not rel_id:
            continue
        part = f"{kind}s/sheet{index}.xml"
        rels.append(f'<Relationship Id="{rel_id}" Target="{part}" '
                    f'Type="{_REL}/{kind}"/>')
        entries.append((f"xl/{part}",
                        f'<?xml version="1.0"?><worksheet xmlns="{_NS}">'
                        f'{body}</worksheet>'))

    listed = "".join(
        f'<sheet name="{name}" sheetId="{i}" r:id="{rel_id}"/>'
        for i, (name, rel_id, _, _) in enumerate(sheets, start=1))

    entries.append(("xl/workbook.xml",
                    f'<?xml version="1.0"?><workbook xmlns="{_NS}" '
                    f'xmlns:r="{_REL}"><sheets>{listed}</sheets></workbook>'))
    entries.append(("xl/_rels/workbook.xml.rels",
                    f'<?xml version="1.0"?><Relationships xmlns="{_PKG}">'
                    f'{"".join(rels)}</Relationships>'))
    if shared is not None:
        items = "".join(f"<si><t>{text}</t></si>" for text in shared)
        entries.append(("xl/sharedStrings.xml",
                        f'<?xml version="1.0"?><sst xmlns="{_NS}" '
                        f'count="{len(shared)}">{items}</sst>'))
    if styles is not None:
        applied = "".join(f'<xf numFmtId="{n}"/>' for n in styles)
        entries.append(("xl/styles.xml",
                        f'<?xml version="1.0"?><styleSheet xmlns="{_NS}">'
                        f'<cellXfs count="{len(styles)}">{applied}</cellXfs>'
                        f'<dxfs count="0"/></styleSheet>'))

    with zipfile.ZipFile(path, "w") as archive:
        for name, text in entries:
            archive.writestr(name, text)
    return str(path)


def _rows(*cells):
    return f'<row r="1">{"".join(cells)}</row>'


def test_a_shared_string_is_looked_up(tmp_path):
    """What Excel actually writes: the text lives in a separate table."""
    path = _handmade_xlsx(
        tmp_path / "f.xlsx",
        [("Device data", "rId1", "worksheet",
          f'<sheetData><row r="18"><c r="K18" t="s"><v>1</v></c></row>'
          f'</sheetData>')],
        shared=["not this one", "6061439WX0"])
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": "6061439WX0"}


def test_a_self_closing_cell_does_not_swallow_the_next_one(tmp_path):
    """A styled-but-empty cell must read blank, not steal its neighbour.

    ``<c r="E18" s="168"/>`` is how Excel writes a formatted empty cell, and
    matching it loosely runs straight past the "/" and on to the *next* cell's
    "</c>" — so E18 silently returned F18's value.
    """
    path = _handmade_xlsx(
        tmp_path / "f.xlsx",
        [("Device data", "rId1", "worksheet",
          '<sheetData><row r="18"><c r="E18" s="168"/>'
          '<c r="F18" t="s"><v>0</v></c></row></sheetData>')],
        shared=["the neighbour"])
    record = calist.read_record(path, {"Model": "E18", "S.N": "F18"})
    assert record == {"Model": "", "S.N": "the neighbour"}


def test_a_vba_module_entry_does_not_shift_the_sheet_order(tmp_path):
    """Older .xlsm files list their macro modules as sheets with no r:id."""
    path = _handmade_xlsx(tmp_path / "f.xlsm", [
        ("MainModule", "", None, ""),
        ("Device data", "rId1", "worksheet",
         '<sheetData><row r="18"><c r="K18" t="str"><v>XR-77341</v></c>'
         '</row></sheetData>'),
    ])
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": "XR-77341"}


def test_a_leading_chartsheet_is_not_counted_as_a_sheet(tmp_path):
    """Chartsheets are not worksheets; openpyxl keeps them apart and so do we.

    Asserted on the sheet list rather than on a value, because a chartsheet
    part holds no cells and would be skipped as empty either way — the list is
    where dropping the rule would actually show, and it is what --inspect
    prints.
    """
    path = _handmade_xlsx(tmp_path / "f.xlsx", [
        ("Trend", "rId1", "chartsheet", ""),
        ("Device data", "rId2", "worksheet",
         '<sheetData><row r="18"><c r="K18" t="str"><v>XR-77341</v></c>'
         '</row></sheetData>'),
    ])
    source = calist._open_workbook(path)
    try:
        assert source.sheet_names == ["Device data"]
        assert source.sheet_name == "Device data"
    finally:
        source.close()
    assert calist.read_record(path, {"S.N": "K18"}) == {"S.N": "XR-77341"}


def test_a_dialogsheet_is_still_a_candidate_sheet(tmp_path):
    """The X-ray workbook opens on one, and it must be skipped for emptiness
    rather than by type — openpyxl counts dialogsheets among its worksheets."""
    path = _handmade_xlsx(tmp_path / "f.xlsm", [
        ("Waveform Dialog", "rId1", "dialogsheet",
         '<sheetData><row r="1"><c r="B1" t="str"><v>stub</v></c></row>'
         '</sheetData>'),
        ("Data entry", "rId2", "worksheet",
         '<sheetData><row r="18"><c r="K18" t="str"><v>never reached</v></c>'
         '</row></sheetData>'),
    ])
    assert calist.read_record(path, {"S.N": "B1"}) == {"S.N": "stub"}


def test_a_date_formatted_number_becomes_a_date(tmp_path):
    """The styles table is only opened for a numeric cell a map asks for, so
    this pins that the lazy read still finds the number format."""
    path = _handmade_xlsx(
        tmp_path / "f.xlsx",
        [("Device data", "rId1", "worksheet",
          '<sheetData><row r="16">'
          '<c r="E16" s="1"><v>45306</v></c>'
          '<c r="F16" s="0"><v>45306</v></c>'
          '</row></sheetData>')],
        styles=[0, 14])          # 14 is the built-in short date format
    record = calist.read_record(path, {"Date": "E16", "S.N": "F16"})
    assert record == {"Date": "2024-01-15 00:00:00", "S.N": "45306"}


def test_escaped_text_is_decoded(tmp_path):
    path = _handmade_xlsx(
        tmp_path / "f.xlsx",
        [("Device data", "rId1", "worksheet",
          '<sheetData><row r="18"><c r="K18" t="str">'
          '<v>Smith &amp; Sons &lt;GmbH&gt; #40</v></c></row></sheetData>')])
    assert calist.read_record(path, {"S.N": "K18"}) == {
        "S.N": "Smith & Sons <GmbH> #40"}


def test_an_unreadable_file_raises_rather_than_reading_blank(tmp_path):
    """A silent blank field is the failure mode --inspect exists to hunt, so
    anything the reader cannot make sense of has to be loud."""
    path = tmp_path / "G302-BB001-0526.xlsx"
    path.write_bytes(b"this is not a zip")
    with pytest.raises(Exception):
        calist.read_record(str(path), {"S.N": "K18"})



# ── finding forms in a folder ─────────────────────────────────────────────────
#
# The intake used to walk with rglob("*") and stat every entry, on the UI
# thread. It also picked up the "~$" lock files Excel leaves beside any open
# workbook, which are not workbooks at all — so a folder someone was working in
# produced a row of failures for files nobody chose.

def _tree(root, names):
    for name in names:
        path = root / name
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"x")
    return root


def test_lock_files_are_not_forms():
    assert not calist.is_source_file("~$G302-AGH001-0425.xlsx")
    assert calist.is_source_file("G302-AGH001-0425.xlsx")


def test_a_previous_run_output_is_not_a_form():
    """The register lands among the sources, so selecting the folder twice
    would otherwise feed the last run's output back in as an input."""
    assert not calist.is_source_file(calist.OUTPUT_NAME)
    assert not calist.is_source_file(calist.OUTPUT_NAME.upper())


def test_only_excel_files_are_found(tmp_path):
    _tree(tmp_path, ["a.xlsx", "b.xls", "c.xlsm", "notes.txt", "sheet.csv"])
    found = {Path(p).name for p in calist.find_source_files(tmp_path)}
    assert found == {"a.xlsx", "b.xls", "c.xlsm"}


def test_the_scan_reaches_sub_folders(tmp_path):
    _tree(tmp_path, ["top.xlsx", "ward/one.xlsx", "ward/lab/two.xlsx"])
    found = {Path(p).name for p in calist.find_source_files(tmp_path)}
    assert found == {"top.xlsx", "one.xlsx", "two.xlsx"}


def test_the_scan_skips_lock_files(tmp_path):
    _tree(tmp_path, ["G302-AGH001-0425.xlsx", "~$G302-AGH001-0425.xlsx"])
    found = [Path(p).name for p in calist.find_source_files(tmp_path)]
    assert found == ["G302-AGH001-0425.xlsx"]


def test_the_scan_stops_when_cancelled(tmp_path):
    _tree(tmp_path, [f"f{n}.xlsx" for n in range(20)])
    cancel = threading.Event()
    cancel.set()
    assert list(calist.find_source_files(tmp_path, cancel)) == []


def test_an_unreadable_folder_does_not_lose_the_rest(tmp_path, monkeypatch):
    """One permission error deep in a tree must not cost the whole round."""
    _tree(tmp_path, ["good.xlsx", "locked/hidden.xlsx"])
    real = os.scandir

    def refuse(path):
        if Path(path).name == "locked":
            raise PermissionError(13, "Access is denied")
        return real(path)

    monkeypatch.setattr(os, "scandir", refuse)
    found = [Path(p).name for p in calist.find_source_files(tmp_path)]
    assert found == ["good.xlsx"]

# ── end-to-end ────────────────────────────────────────────────────────────────

def _form(path, cells):
    wb = Workbook()
    ws = wb.active
    for ref, value in cells.items():
        ws[ref] = value
    wb.save(path)
    return str(path)


@pytest.fixture
def workspace(tmp_path):
    """A source folder, a template, and three forms (one a two-row device)."""
    src = tmp_path / "src"
    src.mkdir()
    tpl_dir = tmp_path / "tpl"
    tpl_dir.mkdir()

    template = tpl_dir / "Device List.xlsx"
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(["No."] + calist.FIELDS, start=1):
        ws.cell(row=3, column=col, value=name)
    wb.save(template)

    files = [
        _form(src / "Clinic-AC001.xlsx",
              {"E17": "Zoll", "E15": "R-Series", "K15": "SN-1",
               "K17": "ER", "E13": "01/01/2024", "G24": "Working"}),
        _form(src / "Clinic-AGH002.xlsx",
              {"E20": "Philips", "E18": "MX450", "K18": "SN-2", "K20": "ICU",
               "E16": "02/01/2024", "D39": "Working", "J39": "Faulty"}),
        _form(src / "Clinic-ZZZ003.xlsx", {"A1": "unknown"}),
    ]
    return files, str(template)



# ── quiet mode and the structured duplicate report ────────────────────────────
#
# Turbo runs tens of thousands of forms, where a line per file makes the log
# useless and the duplicate warnings are the thing worth keeping.

def test_duplicates_are_reported_as_data_not_only_as_a_log_line():
    records = [_sourced("ECG", "SN1", "D23-AGH001-0225.xlsx"),
               _sourced("ECG", "SN1", "D23-AGH007-0225.xlsx")]
    dropped = []
    calist.deduplicate_records(records, dropped)

    assert len(dropped) == 1
    hit = dropped[0]
    assert hit.serial == "SN1"
    assert hit.dropped == "D23-AGH007-0225.xlsx"
    assert hit.kept == "D23-AGH001-0225.xlsx"


def test_collecting_duplicates_is_optional():
    """Callers that only want the log must not have to pass a list."""
    records = [_sourced("ECG", "SN1", "a.xlsx"), _sourced("ECG", "SN1", "b.xlsx")]
    assert len(calist.deduplicate_records(records)) == 1


def test_quiet_mode_drops_the_line_per_file(workspace, caplog):
    files, template = workspace
    with caplog.at_level(logging.INFO, logger="aggregator"):
        calist.process_files(files, template, quiet=True)
    assert "[OK]" not in caplog.text


def test_quiet_mode_still_reports_every_problem(workspace, caplog):
    """It is the successes that make a log unreadable, never the failures."""
    files, template = workspace
    with caplog.at_level(logging.INFO, logger="aggregator"):
        calist.process_files(files, template, quiet=True)
    assert "ZZZ" in caplog.text                  # the unknown-code form


def test_a_loud_run_still_names_every_file(workspace, caplog):
    files, template = workspace
    with caplog.at_level(logging.INFO, logger="aggregator"):
        calist.process_files(files, template)
    assert "[OK]" in caplog.text


def test_quiet_mode_writes_the_same_register(workspace, tmp_path):
    """Logging less must not change a single row."""
    files, template = workspace
    (tmp_path / "loud").mkdir()
    (tmp_path / "quiet").mkdir()

    loud = calist.process_files(files, template, output_dir=tmp_path / "loud")
    quiet = calist.process_files(files, template, output_dir=tmp_path / "quiet",
                                 quiet=True)

    assert loud.succeeded and quiet.succeeded

    assert loud.rows_written == quiet.rows_written
    rows = [[c.value for c in row]
            for path in (loud.output_path, quiet.output_path)
            for row in load_workbook(path).active.iter_rows(min_row=4, max_row=6)]
    assert rows[:3] == rows[3:]


def test_run_reports_every_file_and_writes_the_register(workspace):
    files, template = workspace
    result = calist.process_files(files, template)

    assert result.succeeded
    assert result.output_path.exists()
    assert len(result.outcomes) == 3
    assert result.files_read == 2                 # the ZZZ file is skipped
    assert result.second_rows_added == 1          # AGH generates its NIBP row
    assert result.rows_written == 3               # 2 devices + 1 module row
    assert [o.status for o in result.problems] == [calist.UNKNOWN_CODE]


def test_progress_hook_fires_once_per_file_in_order(workspace):
    files, template = workspace
    seen = []
    calist.process_files(files, template,
                         on_file=lambda o, i, t: seen.append((i, t, o.status)))

    assert [i for i, _, _ in seen] == [1, 2, 3]
    assert {t for _, t, _ in seen} == {3}


def test_cancelling_before_the_run_writes_nothing(workspace):
    files, template = workspace
    cancel = threading.Event()
    cancel.set()

    result = calist.process_files(files, template, cancel=cancel)

    assert result.cancelled
    assert result.output_path is None
    assert not result.succeeded
    assert all(o.status == calist.CANCELLED for o in result.outcomes)


def test_cancelling_partway_stops_early(workspace):
    files, template = workspace
    cancel = threading.Event()

    # Trip the flag as soon as the first file has been handled.
    def on_file(outcome, index, total):
        if index == 1:
            cancel.set()

    result = calist.process_files(files, template, on_file=on_file, cancel=cancel)

    assert result.cancelled
    assert result.output_path is None
    assert result.files_read == 1
    assert any(o.status == calist.CANCELLED for o in result.outcomes)


def test_refuses_to_overwrite_the_template(tmp_path):
    """The template living beside the sources must not be clobbered."""
    src = tmp_path / "src"
    src.mkdir()
    template = src / calist.OUTPUT_NAME       # same folder, same name
    wb = Workbook()
    wb.save(template)
    form_path = _form(src / "Clinic-AC001.xlsx", {"E15": "R-Series"})

    result = calist.process_files([form_path], str(template))

    assert not result.succeeded
    assert "template" in (result.error or "").lower()
    assert template.stat().st_size > 0


def test_strict_names_skips_badly_named_files_in_a_run(tmp_path):
    src = tmp_path / "src"
    src.mkdir()
    tpl_dir = tmp_path / "tpl"
    tpl_dir.mkdir()
    template = tpl_dir / "Device List.xlsx"
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(["No."] + calist.FIELDS, start=1):
        ws.cell(row=3, column=col, value=name)
    wb.save(template)

    good = _form(src / "G302-AC001-0425.xlsx",
                 {"E17": "Zoll", "E15": "R-Series", "K15": "SN-1",
                  "K17": "ER", "E13": "01/01/2024", "G24": "Working"})
    bad = _form(src / "Clinic-AC002.xlsx",
                {"E17": "Zoll", "E15": "R-Series", "K15": "SN-2",
                 "K17": "ER", "E13": "01/01/2024", "G24": "Working"})

    loose = calist.process_files([good, bad], str(template))
    assert loose.files_read == 2

    strict = calist.process_files([good, bad], str(template), strict_names=True)
    assert strict.files_read == 1
    assert [o.status for o in strict.problems] == [calist.BAD_FORMAT]
    assert strict.rows_written == 1


def test_second_row_code_survives_the_house_format():
    """G302-AGH001-0425 must become G302-AGCB001-0425, not mangle the rest."""
    parent = {"Code": "G302-AGH001-0425", "Status": "OK", "Status2": "Faulty",
              "_row_order": 0}
    child = calist.build_second_row(
        parent, {"device_name": "NIBP", "code_replace": ("AGH", "AGCB")})
    assert child["Code"] == "G302-AGCB001-0425"


def test_register_is_signed_with_the_author(workspace):
    """Credit must travel with the file, not just live in the app."""
    files, template = workspace
    result = calist.process_files(files, template)

    wb = load_workbook(result.output_path)
    ws = wb.active
    footer_row = calist.TEMPLATE_START_ROW + result.rows_written + 1
    footer = ws.cell(row=footer_row, column=calist.TEMPLATE_START_COL).value
    creator = wb.properties.creator
    wb.close()

    assert footer and calist.AUTHOR_NAME in footer
    assert calist.AUTHOR_EMAIL in footer
    assert "Calist" in footer
    assert calist.AUTHOR_NAME in creator          # File -> Properties in Excel


def test_attribution_sits_clear_of_the_data(workspace):
    """A blank row between the last record and the signature."""
    files, template = workspace
    result = calist.process_files(files, template)

    wb = load_workbook(result.output_path)
    ws = wb.active
    last_data = calist.TEMPLATE_START_ROW + result.rows_written - 1
    gap = ws.cell(row=last_data + 1, column=calist.TEMPLATE_START_COL).value
    wb.close()

    assert gap in (None, ""), "the signature must not touch the last record"


def test_written_rows_land_where_the_template_expects_them(workspace):
    files, template = workspace
    result = calist.process_files(files, template)

    wb = load_workbook(result.output_path)
    ws = wb.active
    first = [ws.cell(row=calist.TEMPLATE_START_ROW, column=c).value
             for c in range(1, len(calist.FIELDS) + 2)]
    wb.close()

    assert first[0] == 1                              # index in column A
    assert first[1] == "Defibrillator"                # Device in column B
    assert "SN-1" in first[4]                         # S.N in column E


# ── reading a form whose map no longer fits ───────────────────────────────────
#
# Measured over the 2025 and 2026 archives: the configured map works for 94.7%
# of 42,826 forms. The rest were re-laid-out between rounds — a uniform shift of
# one to three rows — and read blank with nothing said. These pin the fallback
# chain, and the two traps that make it dangerous to write naively.

_STANDARD = {"Manufacturer": "E20", "Model": "E18", "S.N": "K18",
             "Location": "K20", "Date": "E16", "Status": "H32"}


def _device_form(path, row, *, sheet="Device data", extra=None, before=None):
    """A form in the usual shape, with its field block anchored at `row`."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, cells in (before or []):
        ws = wb.create_sheet(title)
        for ref, value in cells.items():
            ws[ref] = value
    ws = wb.create_sheet(sheet)
    ws[f"C{row}"] = "Model:"
    ws[f"E{row}"] = "Perfusor Space"
    ws[f"I{row}"] = "Serial No.:"
    ws[f"K{row}"] = "148271"
    ws[f"C{row + 2}"] = "Manufacturer:"
    ws[f"E{row + 2}"] = "B.Braun"
    ws[f"I{row + 2}"] = "Location"
    ws[f"K{row + 2}"] = "ICU"
    for ref, value in (extra or {}).items():
        ws[ref] = value
    wb.save(path)
    return str(path)


def test_the_configured_map_wins_when_it_fits(tmp_path):
    path = _device_form(tmp_path / "f.xlsx", 18)
    record, how = calist.read_best(path, {"cells": _STANDARD})
    assert how == "primary"
    assert record["Model"] == "Perfusor Space"


def test_an_alternate_layout_is_tried_next(tmp_path):
    path = _device_form(tmp_path / "f.xlsx", 19)          # shifted one row down
    config = {"cells": _STANDARD,
              "alt_cells": [{**_STANDARD, "Model": "E19", "S.N": "K19",
                             "Manufacturer": "E21", "Location": "K21"}]}
    record, how = calist.read_best(path, config)
    assert how == "alt 1"
    assert record["S.N"] == "148271"


def test_the_printed_labels_rescue_an_unrecorded_shift(tmp_path):
    """No alt_cells for this offset — the form's own labels have to find it."""
    path = _device_form(tmp_path / "f.xlsx", 15)
    record, how = calist.read_best(path, {"cells": _STANDARD})
    assert how == "labels"
    assert record["Model"] == "Perfusor Space"
    assert record["S.N"] == "148271"
    assert record["Manufacturer"] == "B.Braun"
    assert record["Location"] == "ICU"


def test_the_device_block_is_found_on_another_sheet(tmp_path):
    """GC forms put a Report tab in front; BM keeps the details on a cover."""
    path = _device_form(tmp_path / "f.xlsx", 24, sheet="cover page",
                        before=[("Report", {"A1": "Electrical Safety Test",
                                            "A2": "Protective Earth"})])
    record, how = calist.read_best(path, {"cells": _STANDARD})
    assert how == "labels on 'cover page'"
    assert record["Model"] == "Perfusor Space"


def test_the_calibrator_is_never_read_as_the_device(tmp_path):
    """The trap this whole guard exists for.

    A Hemodialysis cover page prints the reference meter ABOVE the device under
    test, so "first Model: label wins" records the calibrator's model and serial
    as the machine's. The headings are the only reliable signal.
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("cover page")
    ws["A15"] = "Calibration Device"
    ws["G15"] = "Model"
    ws["H15"] = "EMIS"                                   # the Fluke
    ws["G18"] = "Serial No."
    ws["H18"] = "9D2749"
    ws["A22"] = "Device information"
    ws["G24"] = "Model"
    ws["H24"] = "AK96"                                   # the dialysis machine
    ws["A26"] = "Manufacturer"
    ws["C26"] = "Gambro"
    ws["G26"] = "Serial No."
    ws["H26"] = "11195"
    path = tmp_path / "f.xlsx"
    wb.save(path)

    record, how = calist.read_best(str(path), {"cells": _STANDARD})

    assert record["Model"] == "AK96", "read the calibrator instead of the device"
    assert record["S.N"] == "11195"
    assert record["Manufacturer"] == "Gambro"
    assert "EMIS" not in record.values()
    assert "9D2749" not in record.values()


def test_a_table_heading_is_not_mistaken_for_a_field(tmp_path):
    """Therapeutic Ultrasound heads a table 'Model | S.N.' at row 11 and puts
    the real fields at row 69. Taking the first match read the model as "S.N.".
    """
    path = _device_form(tmp_path / "f.xlsx", 69,
                        extra={"A11": "Model", "E11": "S.N."})
    record, how = calist.read_best(path, {"cells": _STANDARD})
    assert record["Model"] == "Perfusor Space"
    assert record["S.N"] == "148271"


def test_a_form_of_placeholders_is_not_rescued(tmp_path):
    """Some cover sheets are filled in with 0 throughout. A serial of "0"
    classifies as a placeholder rather than a blank, so without an explicit
    check the fallback "rescues" them into a row of zeroes.
    """
    path = _device_form(tmp_path / "f.xlsx", 15)
    wb = load_workbook(path)
    ws = wb["Device data"]
    for ref in ("E15", "K15", "E17", "K17"):
        ws[ref] = "0"
    wb.save(path)

    _record, how = calist.read_best(str(path), {"cells": _STANDARD})
    assert how == "none"


def test_nothing_readable_reports_none(tmp_path):
    path = _multi_sheet_form(tmp_path / "f.xlsx", [("Sheet", {"A1": "nothing"})])
    _record, how = calist.read_best(path, {"cells": _STANDARD})
    assert how == "none"


def test_plausible_rejects_a_serial_that_is_not_one():
    assert not calist.plausible({"S.N": "ICU", "Model": "MX450"})
    assert not calist.plausible({"S.N": "16-01-2026", "Model": "MX450"})
    assert not calist.plausible({"S.N": "SN1", "Model": ""})
    assert calist.plausible({"S.N": "SN1", "Model": "MX450"})


def test_classify_serial_sorts_the_shapes_apart():
    assert calist.classify_serial("") == "blank"
    assert calist.classify_serial("N.A") == "placeholder"
    assert calist.classify_serial("ICU") == "suspect"
    assert calist.classify_serial("16-01-2026") == "suspect"
    assert calist.classify_serial("BALANCE001") == "assigned"
    assert calist.classify_serial("STX21170332PA") == "real"


# ── the device table after the archive validation ─────────────────────────────

def test_every_cell_reference_is_a_real_a1_reference():
    """Catches a typo in a coordinate, which would otherwise read blank."""
    from openpyxl.utils.cell import coordinate_to_tuple
    for code, config in DEVICE_CONFIGS.items():
        maps = [config["cells"], *config.get("alt_cells", [])]
        for cell_map in maps:
            for field, ref in cell_map.items():
                if not ref:
                    continue                      # deliberately unmapped field
                coordinate_to_tuple(ref)          # raises if malformed
                assert re.fullmatch(r"[A-Z]{1,3}\d{1,4}", ref), f"{code}.{field}={ref}"


def test_mammography_is_deliberately_not_mapped():
    """BD must stay unmapped until a form is found that is actually filled in.

    All 40 BD workbooks in the 2025-2026 archive carry the same header —
    GE / Alpha st / Gona Hospital, survey date 2012 — across many different
    site codes. That is the vendor QC template's boilerplate, not the device.
    A cell map would write the same fabricated manufacturer and model into 40
    register rows and look entirely plausible.
    """
    assert "BD" not in DEVICE_CONFIGS


def test_the_devices_added_from_the_archive_are_all_named():
    """Every code mapped here is one the site's master list actually names."""
    from device_names import DEVICE_NAMES
    added = {"FA", "EQ", "FV", "FM", "FR", "FF", "DE", "FD", "DB", "CZ", "GM",
             "BQ", "EZ", "FU", "FT", "CP", "FW", "DU", "AY", "BW", "BX", "FP",
             "DW", "DS", "GH", "DX", "BC", "CD", "EN"}
    assert added <= set(DEVICE_CONFIGS), added - set(DEVICE_CONFIGS)
    assert added <= set(DEVICE_NAMES), added - set(DEVICE_NAMES)


def test_a_form_with_no_status_field_maps_status_to_nothing():
    """CT, MRI, Dexa and the air mattress end at "Tested By" — no Status box.

    Mapping Status to a borrowed coordinate would read whatever sits there.
    """
    for code in ("BW", "BX", "FP", "DW"):
        assert DEVICE_CONFIGS[code]["cells"]["Status"] == "", code
