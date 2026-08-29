"""Push a folder of inspection forms into MedCal Pro's Firebase backend.

Calist's own output is a spreadsheet for a human to read. This module takes the
same extraction and points it somewhere else: the Firestore collection the
MedCal Pro mobile app queries, so an engineer standing in front of a device can
be told when it was last calibrated.

**It reuses Calist rather than reimplementing it.** ``find_source_files``,
``classify_file``, ``extract_device_code``, ``read_record`` and the 57 cell maps
in ``device_config`` are imported, not copied. If a device layout is corrected
there, this follows automatically.

Three modes, and they are meant to be run in this order:

    python firebase_export.py --dry-run  "D:/MedCal Pro"
    python firebase_export.py --limit 200 --year 2026  "D:/MedCal Pro"
    python firebase_export.py            "D:/MedCal Pro"

``--dry-run`` opens no network connection and writes nothing. It answers the
questions you cannot answer by looking at the folder: how many files actually
parse, which device codes are unrecognised, how many serials are blank or
duplicated, and which site codes appear in filenames but not in the code list.
**Read that report before letting anything write.** On an archive this size the
cost of finding out afterwards is measured in hours.

Credentials
-----------
The service-account key is read from the ``CALIST_FIREBASE_KEY`` environment
variable, which holds a *path*::

    set CALIST_FIREBASE_KEY=C:\\keys\\medcalpro-admin.json

It is deliberately never a command-line default and never a file inside this
repository. **This repository is public.** A service-account key committed here
would hand anyone on the internet write access to four years of calibration
records. ``.gitignore`` blocks the obvious filenames and a pre-commit hook
refuses anything that looks like a key, but neither is a substitute for keeping
the file somewhere else entirely.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from calist import read_best as calist_read_best
from calist import (
    BAD_FORMAT,
    ERROR,
    LOCK_PREFIX,
    READY,
    UNKNOWN_CODE,
    UNSUPPORTED,
    build_second_row,
    classify_file,
    classify_serial,
    clean,
    extract_device_code,
    find_source_files,
    locate_by_labels,
    plausible,
    read_record,
)
from device_config import DEVICE_CONFIGS
from device_names import DEVICE_NAMES

# ──────────────────────────────────────────────────────────────────────────────
# What the app expects
# ──────────────────────────────────────────────────────────────────────────────

#: The Firestore collection the mobile app already reads. Imported history goes
#: into the *same* collection as work filed from the app, not a parallel one,
#: so the previous-calibration lookup and the duplicate guard need no second
#: code path — both already query here.
CALIBRATIONS = "calibrations"

#: One document per site, keyed by customer code.
CUSTOMERS = "customers"

#: One tombstone per calibration an engineer deleted from the app, keyed by the
#: record id it replaces.
#:
#: Deleting the record alone would not stick. This import is idempotent and
#: keyed on the filename, so the next run over the archive recreates whatever it
#: finds on disk — the engineer would watch a device they removed come back, and
#: stop trusting the button. Every write path here reads this collection first
#: and skips what it names.
DELETIONS = "deletions"

#: Cloud Storage prefix. Mirrors the on-disk shape so a storagePath stays
#: legible to somebody reading a Firestore document.
STORAGE_PREFIX = "archive"

#: The app's CustomerCategory enum. Anything the code list calls something else
#: lands in ``other`` rather than being silently forced into ``hospital``.
CATEGORY_HOSPITAL = "hospital"
CATEGORY_CENTER = "center"
CATEGORY_COMPANY = "company"
CATEGORY_AUTHORITY = "healthCareAuthority"
CATEGORY_OTHER = "other"

#: Sheet name in Hospital Codes.xlsx -> the app's category.
#:
#: Every hospital *variant* collapses to ``hospital``: the app draws one badge
#: for all of them, and preserving "University" versus "Insurance" here would
#: promise a distinction nothing downstream can render. The original sheet name
#: is kept on the customer document as ``sourceCategory`` so nothing is lost.
SHEET_CATEGORIES: dict[str, str] = {
    "companies": CATEGORY_COMPANY,
    "centers": CATEGORY_CENTER,
    "insurance hospital": CATEGORY_HOSPITAL,
    "governmental hospitals": CATEGORY_HOSPITAL,
    "university hospitals": CATEGORY_HOSPITAL,
    "education hospitals": CATEGORY_HOSPITAL,
    "specialized hospitals": CATEGORY_HOSPITAL,
    "private hospitals": CATEGORY_HOSPITAL,
}

#: Arabic folder names seen under each year, normalised to a category.
#:
#: The spellings genuinely drift between years — "الهيئة العامه للرعايه الصحية"
#: in 2023-24 became "الهيئة العامة للرعاية الصحية" in 2025-26, and "مراكز"
#: grew an article to become "المراكز". Both spellings are listed rather than
#: normalised by rule, because a rule that strips articles and normalises
#: taa-marbuta would also merge things that should stay apart.
FOLDER_CATEGORIES: dict[str, str] = {
    "مستشفيات": CATEGORY_HOSPITAL,
    "مراكز": CATEGORY_CENTER,
    "المراكز": CATEGORY_CENTER,
    "شركات": CATEGORY_COMPANY,
    "الهيئة العامه للرعايه الصحية": CATEGORY_AUTHORITY,
    "الهيئة العامة للرعاية الصحية": CATEGORY_AUTHORITY,
    # Everything below is a real folder in the archive that is not one of the
    # app's four. They are not errors — they are ministries, tenders and
    # authorities — so they land in ``other`` deliberately.
    "الشراء الموحد": CATEGORY_OTHER,
    "الامانة العامة": CATEGORY_OTHER,
    "الانتاج الحربى": CATEGORY_OTHER,
    "هيئة قناة السويس": CATEGORY_OTHER,
    "وزارات": CATEGORY_OTHER,
    "المناقصة 2025": CATEGORY_OTHER,
    "معهد ناصر": CATEGORY_OTHER,
}

def named_only(code: str) -> str | None:
    """The device a code names, when there is no cell map to read its form.

    Every code in the site's master list that ``device_config`` has no layout
    for. Such a record still carries the right device type, the site, the date
    and the file — most of what an inventory needs — and says plainly that its
    readings could not be extracted.

    **Deliberately not merged into ``device_config``.** A code there without a
    correct cell map would make Calist emit a register row full of whatever
    happens to sit at those coordinates, which is worse than reporting the file
    as unrecognised: a plausible-looking wrong row does not announce itself.

    This began as four codes typed by hand. The master list turned it into
    ~110, which is the difference between a thousand archive files importing as
    "unknown device" and importing as the device they actually are.
    """
    key = code.strip().upper()
    if not key or key in DEVICE_CONFIGS:
        return None
    return DEVICE_NAMES.get(key)

#: Files that are not inspection forms and must never become records.
#:
#: The first dry run only excluded names starting with "device list", and the
#: collision report showed what that missed: `List.xlsx`, `LIST.xlsx`,
#: `Sphyg list.xlsx`, `Copy of Device List.xlsx` — registers under a dozen
#: different names — plus scratch files (`New Microsoft Excel Worksheet`) and
#: per-device working sheets (`OCCLUSION`, `Uncertainity`) that are calculation
#: aids, not calibrations.
#:
#: Importing any of these would add a record per row for an entire site.
_NOT_A_FORM = re.compile(
    r"""^\s*(
          (copy\s+of\s+)?(device\s*)?list        # list, device list, copy of device list
        | .*\blist\b.*                           # sphyg list, device list all, list 2
        | new\s+microsoft\s+excel\s+worksheet
        | occlusion
        | uncertain(i)?ty
        | book\d*
        | template
    )\s*\d*\s*[.\-]*\s*$""",
    re.IGNORECASE | re.VERBOSE,
)

#: SITE-TAG-MMYY, e.g. G302-AGH001-0226. Same shape Calist validates against;
#: restated here because this module needs the *parts*, not a yes/no.
_STEM_RE = re.compile(
    r"^(?P<site>[A-Za-z]+\d+)-(?P<tag>[A-Za-z]+\d+)-(?P<mm>0[1-9]|1[0-2])(?P<yy>\d{2})$"
)

#: A serial assigned in-house when the device carries no manufacturer serial.
#: The house convention is letters, optionally a dot, then digits — and it
#: always ends in digits: BALANCE001, BALANCE 001, N.A002.
_ASSIGNED_SERIAL_RE = re.compile(r"^[A-Za-z][A-Za-z ]*\.?\s*\d+$")

#: Contains a digit anywhere.
#:
#: The test is *any* digit, not a trailing one. An earlier version required the
#: value to end in a digit and duly flagged STX21170332PA, NV03124H and
#: S0QQM3HDC00061J — all perfectly good manufacturer serials that happen to end
#: in a letter. What actually distinguishes a serial from a location typed into
#: the wrong cell is that a serial has numbers in it at all and "ICU" does not.
_HAS_DIGIT_RE = re.compile(r"\d")

#: Written where a device genuinely has no serial and none was assigned.
_PLACEHOLDERS = {"", "-", "--", "n.a", "na", "n/a", "none", "null", "0"}

#: A date, in any of the shapes these forms use.
#:
#: Needed because "has a digit in it" is not enough to identify a serial. When
#: the alternate Infusion layout was tried against a form that did not use it,
#: the serial cell landed on the next-calibration date — "16-01-2026" — which
#: passed a digits-only test and would have been imported as a serial number.
#: The printed label on those forms confirmed the original layout was the right
#: one. Dates are the single most common wrong answer, so they are excluded by
#: name rather than left to chance.
_DATE_LIKE_RE = re.compile(
    r"^\s*\d{1,4}\s*[-/.]\s*\d{1,2}\s*[-/.]\s*\d{1,4}\s*$"
)


# One parsed form
# ──────────────────────────────────────────────────────────────────────────────


@dataclass
class ParsedForm:
    """One inspection form, resolved as far as it can be without opening it."""

    path: str
    filename: str
    year: int

    #: Everything between the category folder and the file. Reported so an
    #: unmapped site code can be identified by hand; never parsed for a name.
    folder_trail: str = ""
    folder_category: str = ""

    #: From the filename. Empty when it does not follow the convention.
    site: str = ""
    tag: str = ""
    month: int = 0

    #: Two-digit year from the filename's MMYY. This is the inspection date and
    #: the identity; ``year`` above is only which folder it was filed in.
    filed_yy: int = 0

    device_code: str = ""
    device_name: str = ""

    #: Read from inside the workbook. Empty until a deep pass runs.
    serial: str = ""
    model: str = ""
    manufacturer: str = ""
    location: str = ""
    form_date: str = ""
    status: str = ""

    #: The client as the *form* names it, read from the header rather than the
    #: code list. Worth having for two reasons: it is the only source of an
    #: address anywhere in the pipeline, and it covers sites the code list has
    #: never heard of — 25 of them, carrying 7,610 calibrations between them.
    client_name: str = ""
    client_address: str = ""

    #: The engineer's contact at the site, and their number. Recorded per visit
    #: in the form itself, which is the right grain -- a returning customer gets
    #: a fresh contact each year.
    contact_name: str = ""
    contact_phone: str = ""

    #: Which cell layout read this form: 0 is the device's primary map, 1+ an
    #: alternate. Reported so a form that changed between rounds is visible
    #: rather than silently handled.
    layout: int = 0

    #: Set when something is wrong but the file is still worth keeping.
    attention: str = ""

    @property
    def conforms(self) -> bool:
        """True when the filename gave us a site and a device tag."""
        return bool(self.site and self.tag)

    @property
    def customer_code(self) -> str:
        return self.site.upper()

    @property
    def tag_number(self) -> int:
        """The numeric part of the tag: AH132 -> 132, or 0 if there is none."""
        match = _TAG_RE.match(self.tag)
        return int(match.group(2)) if match else 0

    #: Final Firestore id, set by [assign_ids] once serials are known. Empty
    #: until then; nothing may write a document before it is filled in.
    doc_id: str = ""

    #: The tag as it appears in the filename, kept when [tag] has been reassigned
    #: because several devices were sharing it. The file on disk is never
    #: renamed, so this is the only way back from a record to its paperwork.
    original_tag: str = ""

    @property
    def base_doc_id(self) -> str:
        """Deterministic Firestore document id.

        **Idempotency is the whole point.** An 88,000-file import will be
        interrupted, resumed and re-run; an id derived from the content means a
        second pass updates the same document instead of creating a twin. For a
        conforming file it is exactly the filename stem, so somebody reading the
        console can match a document to a file on disk without a lookup.

        The year comes from the **filename**, never from the folder. The first
        dry run got this wrong and the collision report caught it:
        `H15-CE001-1222.xlsx` sitting in the `Customers 2023` folder — December
        2022 work filed in the 2023 round — was given the id `H15-CE001-1223`,
        colliding with the genuinely different `H15-CE001-1223.xlsx`. The
        filename is the identity; the folder only says which round it was filed
        in, which is kept separately as ``sourceYear``.

        A non-conforming file falls back to a hash of its path. It must be the
        *path*, not the stem: an archive this size has four different
        `List.xlsx` and four different `Device.xlsx`, and hashing the stem alone
        collapsed them into one document.
        """
        if self.conforms:
            return f"{self.site}-{self.tag}-{self.month:02d}{self.filed_yy:02d}".upper()
        digest = hashlib.sha1(self.path.encode("utf-8", "replace")).hexdigest()[:16]
        return f"x_{self.year}_{digest}"


#: Splits a device tag into its letter prefix and its number: AGH086 -> AGH, 86.
_TAG_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _next_tag(form: ParsedForm, used: dict[tuple[str, str], set[int]]) -> str:
    """The next unused device number for this device type at this site.

    When several devices share one filename, the extras need numbers of their
    own — `AH132` appearing five times at ORASCOM becomes `AH132` plus four
    numbers carrying on from the highest `AH` already in use at that site. That
    is how the numbering works on paper, so it is what an engineer will
    recognise when they meet the device again.

    ``used`` is seeded from every tag in the archive before any assignment, so
    a new number can never land on a device that already exists. Digit width is
    copied from the tag being replaced, so `AGH086` yields `AGH301`, not
    `AGH0301`.
    """
    match = _TAG_RE.match(form.tag)
    if not match:
        return form.tag
    prefix, digits = match.group(1), match.group(2)

    key = (form.customer_code, prefix.upper())
    taken = used.setdefault(key, set())
    candidate = (max(taken) + 1) if taken else (int(digits) + 1)
    while candidate in taken:
        candidate += 1
    taken.add(candidate)

    return f"{prefix}{candidate:0{len(digits)}d}"


def assign_ids(forms: list[ParsedForm]) -> tuple[int, int]:
    """Give every form its final document id. Returns (collapsed, split).

    ``SITE-TAG-MMYY`` is **not unique in this archive**, and treating it as
    unique silently lost 1,305 of 16,038 records on the first 2026 import. Two
    quite different things produce a clash, and they need opposite handling:

      * **The same device, filed twice.** Copies left behind in working
        sub-folders — `PM/` and `PM/24-safwa-pm/`, or a folder called
        `Printeddddddd`. One calibration, several files. These must collapse to
        one record, or the site inventory counts the device several times.

      * **Different devices sharing a tag.** ORASCOM reused `AH132` across five
        branches in one month — Mansoura, Ras El Hekma, the Housing Bank.
        Five real calibrations of five real devices. Collapsing them throws
        four away.

    **The serial tells them apart**, so identity is site + tag + month + serial,
    with the serial only entering the id when it has to. Suffixing by a hash of
    the serial rather than by position keeps the id stable no matter what order
    the files are walked in or where they are moved to, which is what makes a
    re-run update rather than duplicate.
    """
    # Every device number already in use, per site and device type. Seeded from
    # the whole archive *before* anything is reassigned, so a number invented
    # here can never land on a device that genuinely exists somewhere else in
    # the same site's history — in any year.
    used: dict[tuple[str, str], set[int]] = defaultdict(set)
    for form in forms:
        match = _TAG_RE.match(form.tag)
        if match and form.customer_code:
            used[(form.customer_code, match.group(1).upper())].add(
                int(match.group(2))
            )

    # ── Phase 1: settle each device's tag once, across every year ────────
    #
    # A device tag belongs to the *device*, not to the visit: `G302-AGH001`
    # is the same patient monitor whether the file ends -0225 or -0726, and
    # only the date changes. That is what makes "when was this last
    # calibrated" answerable at all.
    #
    # So renumbering cannot be decided per year. If `AH132` covers five
    # devices at ORASCOM, and the same five come back the following year,
    # resolving each year separately would hand the same physical device a
    # different number in each — and the previous-calibration lookup, which
    # matches on site+tag, would never find its own history.
    #
    # The serial is what ties a device to itself across years, so the mapping
    # is built once here, keyed on (site, tag-in-filename, serial), and then
    # applied to every year at once.
    per_device: dict[tuple[str, str], dict[str, list[ParsedForm]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for form in forms:
        if form.tag and form.serial.strip():
            key = (form.customer_code, form.tag.upper())
            per_device[key][form.serial.strip().upper()].append(form)

    renamed = 0
    for (site, tag), by_serial in sorted(per_device.items()):
        if len(by_serial) < 2:
            continue  # tag means one device here; nothing to resolve
        # Lowest serial keeps the tag the filenames carry; the rest are given
        # numbers that are free at this site across the whole archive.
        for position, serial in enumerate(sorted(by_serial)):
            if position == 0:
                continue
            members = by_serial[serial]
            fresh = _next_tag(members[0], used)
            for member in members:
                member.original_tag = tag
                member.tag = fresh
            renamed += 1

    # ── Phase 2: group by what the document id will now be ───────────────
    groups: dict[str, list[ParsedForm]] = defaultdict(list)
    for form in forms:
        groups[form.base_doc_id].append(form)

    collapsed = split = 0
    # Sorted so assignment order does not depend on how the filesystem was
    # walked; without this the same archive could produce different numbers on
    # two runs and duplicate every renumbered record.
    for base in sorted(groups):
        members = groups[base]
        if len(members) == 1:
            members[0].doc_id = base
            continue

        by_serial: dict[str, list[ParsedForm]] = defaultdict(list)
        for member in members:
            by_serial[member.serial.strip().upper()].append(member)

        # A single *known* serial means one device filed several times.
        #
        # A single *empty* serial means nothing of the sort. Forms whose device
        # code is unmapped are never opened, so their serial is unknown rather
        # than matching — collapsing those would quietly discard real, distinct
        # calibrations on the strength of a comparison that was never made.
        # They fall through to path-based disambiguation instead.
        if len(by_serial) == 1 and next(iter(by_serial)):
            for member in members:
                member.doc_id = base
            collapsed += len(members) - 1
            continue

        # Genuinely different devices behind one filename — or forms we could
        # not read well enough to claim otherwise. Each gets a real, unused
        # device number at its site; see [_next_tag].
        #
        # Sorted so the assignment is the same on every run: the first serial
        # keeps the tag the filename actually carries, and only the extras are
        # renumbered. Renumbering all of them would detach every one of these
        # records from its paperwork for no gain.
        # Known serials first and in order, unknowns last, so the assignment is
        # identical on every run.
        ordered = sorted(by_serial.items(), key=lambda kv: (kv[0] == "", kv[0]))

        # Exactly one record keeps the tag the filename actually carries. The
        # rest are renumbered, and each remembers what it was called so the
        # paperwork can still be found — the files themselves are never renamed.
        keeps_original = True
        created = 0

        def renumber(members: list[ParsedForm]) -> None:
            """Give one device a fresh number, shared by all its copies."""
            nonlocal keeps_original
            if keeps_original:
                keeps_original = False
            else:
                was = members[0].tag
                # Read the prefix and width from the tag before overwriting it.
                fresh = _next_tag(members[0], used)
                for member in members:
                    member.original_tag = was
                    member.tag = fresh
            for member in members:
                member.doc_id = member.base_doc_id

        for serial, sharing in ordered:
            if serial:
                # One device, however many copies of its paperwork exist.
                renumber(sharing)
                created += 1
            else:
                # Serial unknown, so nothing says these are the same device.
                # Each is treated as its own — a spurious extra record is
                # visible and fixable, a silently merged one is neither.
                for member in sharing:
                    renumber([member])
                    created += 1

        split += created - 1

    return collapsed, split


# ──────────────────────────────────────────────────────────────────────────────
# Walking the archive
# ──────────────────────────────────────────────────────────────────────────────


def year_folders(root: Path) -> list[tuple[int, Path]]:
    """Every ``Customers <year>`` folder under ``root``, oldest first."""
    found: list[tuple[int, Path]] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        match = re.search(r"(20\d{2})", child.name)
        if match and child.name.lower().startswith("customers"):
            found.append((int(match.group(1)), child))
    return found


def is_output_file(name: str) -> bool:
    """True for registers, scratch files and working sheets — never forms.

    Deliberately conservative in one direction only: a file that *is* a form and
    happens to be called "list" would be skipped, but no real form in this
    archive is named that way, and the opposite mistake — importing a register —
    silently multiplies a whole site's device count.
    """
    return bool(_NOT_A_FORM.match(Path(name).stem))


def locate(path: Path, year_root: Path) -> tuple[str, str]:
    """Resolve (folder trail, category) from where a file sits.

    **The category is reliable; the customer is not.** The first component below
    the year is the category folder and that is trustworthy. What sits *below*
    it is not: the plain trees really are Category / Customer / … , but the
    Health Care Authority tree — 63,000 files, 72% of the archive — nests
    Category / Governorate / Month / Hospital / device / file, and also holds
    `Reports`, `training` and `BLANK الهيئة` folders that are not customers at
    all.

    The first dry run assumed depth 2 was always the customer and duly reported
    "جنوب سيناء" (a governorate) as one. So this deliberately no longer guesses:
    it returns the whole trail for a human to read in the report, and the
    customer's *identity* comes from the site code in the filename, which is
    unambiguous. Codes absent from the code list are listed for manual mapping
    rather than being given a fabricated name.
    """
    try:
        parts = path.relative_to(year_root).parts
    except ValueError:
        return "", ""

    if not parts:
        return "", ""

    category = FOLDER_CATEGORIES.get(parts[0].strip(), CATEGORY_OTHER)
    # Everything between the category and the file, which is where the customer
    # is — at an unpredictable depth.
    trail = "/".join(p.strip() for p in parts[1:-1])
    return trail, category


def scan(root: Path, only_year: int | None = None, limit: int = 0) -> list[ParsedForm]:
    """Classify every form under ``root`` without opening any of them.

    ``classify_file`` does no workbook I/O, so this is fast enough to run over
    the whole archive — which is what makes a meaningful dry run possible at
    all.
    """
    forms: list[ParsedForm] = []

    for year, folder in year_folders(root):
        if only_year and year != only_year:
            continue

        for filepath in find_source_files(folder):
            name = Path(filepath).name
            if name.startswith(LOCK_PREFIX) or is_output_file(name):
                continue

            path = Path(filepath)
            trail, category = locate(path, folder)
            form = ParsedForm(
                path=filepath,
                filename=name,
                year=year,
                folder_trail=trail,
                folder_category=category,
            )

            # Trailing spaces before the extension are common ("…-0824   .xlsx")
            # and would otherwise fail the pattern for no good reason.
            stem_match = _STEM_RE.match(path.stem.strip())
            if stem_match:
                form.site = stem_match.group("site").upper()
                form.tag = stem_match.group("tag").upper()
                form.month = int(stem_match.group("mm"))
                form.filed_yy = int(stem_match.group("yy"))
                named_year = 2000 + form.filed_yy

                # A file dated December 2022 living in the 2023 folder is
                # normal — it is work filed in the 2023 round. A file dated
                # January 2026 in the 2025 folder is a round that ran over the
                # new year. Both are real, and both are simply noted.
                #
                # A file dated *more than a year after* its round is not real.
                # The archive holds names ending 2032, 2033 and 2055, which are
                # plainly mistyped: a calibration cannot be filed years before
                # it happened. The folder year is the trustworthy half, so the
                # date falls back to it and the original is recorded.
                if named_year > year + 1:
                    form.attention = (
                        f"filename says {form.month:02d}/{named_year}, which is "
                        f"after the {year} round it was filed in — treated as "
                        f"a typo for {year}"
                    )
                    form.filed_yy = year % 100
                elif named_year != year:
                    form.attention = (
                        f"dated {form.month:02d}/{named_year}, "
                        f"filed in the {year} round"
                    )

            outcome = classify_file(filepath)
            if outcome.status == READY:
                form.device_code = outcome.device_code or ""
                form.device_name = DEVICE_CONFIGS[form.device_code]["device_name"]
            elif named_only(extract_device_code(name) or ""):
                # Identified, but not readable. The record still carries the
                # right device type, the file and the site — which is most of
                # what the inventory needs — and says plainly that its readings
                # could not be extracted.
                form.device_code = (extract_device_code(name) or "").upper()
                form.device_name = named_only(form.device_code) or ""
                form.attention = "device known, form layout not mapped yet"
            else:
                form.device_code = (extract_device_code(name) or "").upper()
                reasons = {
                    UNKNOWN_CODE: f"device code '{form.device_code or '?'}' not in the table",
                    BAD_FORMAT: "filename does not follow SITE-TAG-MMYY",
                    UNSUPPORTED: "unsupported file type",
                    ERROR: "could not be classified",
                }
                form.attention = reasons.get(outcome.status, outcome.detail)

            forms.append(form)
            if limit and len(forms) >= limit:
                return forms

    return forms


#: Where a form's header fields can sit.
#:
#: Read as one block and searched **by label**, never by fixed cell. The client
#: name sits at C3 on the Defibrillator sheet, C5 on Anaesthesia, C6 on the
#: Patient Monitor, C11 on the Nebulizer, D6 on the Baby Incubator; the contact
#: is at A30, G32, F44, H42, B28. Nothing about the position is stable.
#:
#: The *labels* are. "Client Name:", "Client Address:", "Contact Person Name:"
#: and "Phone No.:" are word-for-word identical across every device type in the
#: archive, so one rule covers all 57 maps without touching any of them.
#: Contacts reach row 45 and often sit in column A, which is why the block is
#: this large.
_HEADER_COLS = "ABCDEFGHIJKL"
_HEADER_BLOCK = {
    f"{col}{row}": f"{col}{row}"
    for row in range(1, 56)
    for col in _HEADER_COLS
}

#: Anything that is a label rather than a value. Used to step over a label's own
#: merged span — Calist resolves merges, so the cells beside "Client Name:"
#: report that same text, and the first non-empty neighbour is not the answer.
_ANY_LABEL = re.compile(
    r"^\s*(client\s*(name|address)|contact\s*person\s*name|phone\s*(no|number)?)"
    r"[\s:.]*$",
    re.I,
)

_LABELS = {
    # Trailing punctuation is [\s:.]* rather than a single optional character:
    # the label is written "Phone No.:" -- a full stop AND a colon -- which a
    # one-character class silently fails to match.
    "client_name": re.compile(r"^\s*client\s*name[\s:.]*$", re.I),
    "client_address": re.compile(r"^\s*client\s*address[\s:.]*$", re.I),
    "contact_name": re.compile(r"^\s*contact\s*person\s*name[\s:.]*$", re.I),
    "contact_phone": re.compile(r"^\s*phone\s*(no|number)?[\s:.]*$", re.I),
}


def header_from_grid(grid: dict) -> dict:
    """Client and contact details, located by label.

    Returns a dict of the four fields, each empty when its label is absent —
    some device types (`CE` among them) put their test data on the sheet the
    reader picks and the header on another, so nothing is found and the caller
    falls back to the code list.

    Phone numbers keep their leading zero: they are read as text and never
    coerced to a number, which is the same reason the Excel writer uses
    ``TextCellValue`` for them.
    """
    text = {
        ref: (str(value).strip() if value is not None else "")
        for ref, value in grid.items()
    }

    def value_beside(ref: str) -> str:
        match = re.match(r"([A-Z]+)(\d+)", ref)
        if not match:
            return ""
        column, row = match.group(1), match.group(2)
        for candidate in _HEADER_COLS[_HEADER_COLS.index(column) + 1:]:
            found = text.get(f"{candidate}{row}", "")
            if found and not _ANY_LABEL.match(found):
                return found
        return ""

    found: dict[str, str] = {key: "" for key in _LABELS}
    for ref, value in text.items():
        if not value:
            continue
        for key, pattern in _LABELS.items():
            if not found[key] and pattern.match(value):
                found[key] = value_beside(ref)

    return found


def read_best(path: str, config: dict) -> tuple[dict, int]:
    """Read a form, trying every layout Calist knows. See calist.read_best.

    Kept as a wrapper for the integer contract this module's callers use: 0 for
    the configured map, 1+ for an alternate, **-1 when nothing produced a
    plausible record**. That last case matters — an earlier version returned the
    primary's output regardless, so a form on an unknown layout was imported
    with whatever happened to sit at those coordinates and nothing said so.

    The client header is read in the *same* pass as the mapped cells: opening
    each workbook twice would double a 35-minute read for the sake of two
    strings, and the reader resolves a whole block of refs in one go.
    """
    try:
        record, how = calist_read_best(path, config, extra=_HEADER_BLOCK)
    except Exception:  # noqa: BLE001
        return {}, -1
    if how == "primary":
        return record, 0
    if how == "none":
        return record, -1
    if how.startswith("alt "):
        return record, int(how.split()[1])
    # Located from the form's own printed labels — not one of the written-down
    # layouts, so reported as an alternate rather than as the primary.
    return record, len(config.get("alt_cells", [])) + 1


def deepen(forms: list[ParsedForm], sample: int = 0, progress=None) -> int:
    """Open workbooks and fill in serial, model, manufacturer, location, date.

    Returns how many were actually read. This is the expensive half — every
    call opens a file — so the dry run samples by default and the real import
    reads everything.
    """
    targets = [f for f in forms if f.device_code in DEVICE_CONFIGS]
    if sample:
        # Evenly spaced rather than the first N: the archive is ordered by year
        # and site, so the first N would all come from one hospital in 2023 and
        # tell you nothing about the rest.
        step = max(1, len(targets) // sample)
        targets = targets[::step][:sample]

    read = 0
    for index, form in enumerate(targets):
        config = DEVICE_CONFIGS[form.device_code]
        try:
            record, layout = read_best(form.path, config)
            form.serial = clean(record.get("S.N"))
            form.model = clean(record.get("Model"))
            form.manufacturer = clean(record.get("Manufacturer"))
            form.location = clean(record.get("Location"))
            form.form_date = clean(record.get("Date"))
            form.status = clean(record.get("Status"))
            header = header_from_grid(record)
            form.client_name = header["client_name"]
            form.client_address = header["client_address"]
            form.contact_name = header["contact_name"]
            form.contact_phone = header["contact_phone"]
            form.layout = layout
            if layout < 0 and not form.attention:
                form.attention = (
                    "no known layout reads this form — values may be wrong"
                )
            read += 1
        except Exception as error:  # noqa: BLE001 - one bad file must not stop the run
            form.attention = f"could not be read: {type(error).__name__}"
        if progress:
            progress(index + 1, len(targets))

    return read


# ──────────────────────────────────────────────────────────────────────────────
# The customer code list
# ──────────────────────────────────────────────────────────────────────────────


def load_code_list(path: Path) -> dict[str, dict]:
    """Read Hospital Codes.xlsx into ``{CODE: {name, category, sourceCategory}}``.

    The sheet is laid out for a human: a Category column that is filled in only
    on the first row of each block, blank spacer rows throughout, and one sheet
    per kind of customer. Category is carried forward rather than required on
    every row, which is what the blank cells mean.
    """
    import openpyxl  # local: the dry run should not pay for this unless asked

    codes: dict[str, dict] = {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for sheet_name in workbook.sheetnames:
        category = SHEET_CATEGORIES.get(sheet_name.strip().lower(), CATEGORY_OTHER)
        sheet = workbook[sheet_name]
        carried = sheet_name.strip()

        for row in sheet.iter_rows(values_only=True):
            cells = [clean(c) for c in row]
            if not any(cells):
                continue

            # Find a code-shaped cell and take the next non-empty one as the
            # name. Column positions differ between sheets, so anchoring on the
            # code is more robust than trusting an index.
            code = name = ""
            for index, value in enumerate(cells):
                if re.fullmatch(r"[A-Za-z]{1,3}\d{1,4}", value or ""):
                    code = value.upper()
                    name = next((c for c in cells[index + 1 :] if c), "")
                    break

            if not code:
                # A row with text but no code is either the block's category
                # heading or the sheet title.
                label = next((c for c in cells if c), "")
                if label and label.lower() not in ("hospital codes", "category", "code"):
                    carried = label
                continue

            codes[code] = {
                "name": name,
                "category": category,
                "sourceCategory": carried,
                "sheet": sheet_name,
            }

    workbook.close()
    return codes


# ──────────────────────────────────────────────────────────────────────────────
# Writing to Firebase
# ──────────────────────────────────────────────────────────────────────────────

#: Firestore's hard limit on operations in one batched write.
BATCH_SIZE = 500

#: Environment variable holding the *path* to the service-account key.
#:
#: A path, never the key itself, and never a default inside this repository.
#: This repository is public: a key committed here would grant anyone on the
#: internet write access to every calibration record, bypassing the security
#: rules entirely.
KEY_ENV = "CALIST_FIREBASE_KEY"


def using_emulator() -> bool:
    """True when the Firestore emulator is the target rather than the project.

    `firebase-admin` honours FIRESTORE_EMULATOR_HOST on its own; this exists so
    the tool can *say* which one it is about to write to. Writing to production
    when you believed you were on the emulator is the expensive mistake, and it
    is silent unless something announces it.
    """
    return bool(os.environ.get("FIRESTORE_EMULATOR_HOST", "").strip())


def connect(project: str):
    """Open an admin connection, or explain exactly what is missing.

    Imported lazily so that ``--dry-run`` and ``--verify-maps`` — the modes that
    matter most and touch no network — keep working on a machine that has never
    installed firebase-admin.

    Against the emulator no credentials are needed at all, which is a feature
    rather than a shortcut: experimenting should not require handling the key
    that grants write access to four years of real records.
    """
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore, storage
    except ImportError:
        raise SystemExit("firebase-admin is not installed.  pip install firebase-admin")

    if using_emulator():
        # google-cloud-storage reads its own variable, not the one firebase
        # tooling sets. Without this the uploads go to the real bucket while
        # everything else goes to the emulator — the worst possible split.
        gcs_host = os.environ.get("FIREBASE_STORAGE_EMULATOR_HOST", "").strip()
        if gcs_host and not os.environ.get("STORAGE_EMULATOR_HOST"):
            os.environ["STORAGE_EMULATOR_HOST"] = f"http://{gcs_host}"

        if not firebase_admin._apps:
            import google.auth.credentials

            class _EmulatorToken(google.auth.credentials.Credentials):
                """A token the emulator will accept and never has to refresh.

                ``AnonymousCredentials`` looks like the obvious choice and works
                for Firestore, but Cloud Storage asks its credentials to refresh
                and anonymous ones raise on that — so uploads failed with
                "Anonymous credentials cannot be refreshed" while every other
                call succeeded. The emulator does not check the value, it just
                needs one that exists.
                """

                def __init__(self) -> None:
                    super().__init__()
                    self.token = "owner"

                def refresh(self, request) -> None:  # noqa: D401, ARG002
                    self.token = "owner"

                @property
                def expired(self) -> bool:
                    return False

                @property
                def valid(self) -> bool:
                    return True

            class _EmulatorCredential(credentials.Base):
                """No real credentials — the emulator does not check them.

                Without this, ``initialize_app()`` with no credential falls back
                to Application Default Credentials and fails on a machine that
                has never run `gcloud auth`. The whole point of the emulator is
                that experimenting should not require the admin key, so it has
                to work with nothing configured.
                """

                def get_credential(self):
                    return _EmulatorToken()

            firebase_admin.initialize_app(
                _EmulatorCredential(),
                {
                    "projectId": project,
                    "storageBucket": f"{project}.firebasestorage.app",
                },
            )
        return firestore.client(), storage

    key_path = os.environ.get(KEY_ENV, "").strip().strip('"')
    if not key_path:
        raise SystemExit(
            f"{KEY_ENV} is not set.\n"
            f"  Point it at your service-account JSON, kept OUTSIDE this repo:\n"
            f'    setx {KEY_ENV} "C:\\keys\\medcalpro-admin.json"\n'
            f"  Then open a new terminal so the variable is visible.\n\n"
            f"  Or work against the emulator instead, which needs no key and\n"
            f"  costs nothing:\n"
            f"    firebase emulators:start --only firestore,auth,storage\n"
            f"    set FIRESTORE_EMULATOR_HOST=localhost:8080"
        )
    if not Path(key_path).is_file():
        raise SystemExit(f"{KEY_ENV} points at {key_path}, which does not exist.")

    if not firebase_admin._apps:
        firebase_admin.initialize_app(
            credentials.Certificate(key_path),
            {"storageBucket": f"{project}.firebasestorage.app"},
        )
    return firestore.client(), storage


def push_firestore(
    forms: list[ParsedForm],
    codes: dict[str, dict],
    db,
    *,
    limit: int = 0,
    progress=None,
) -> tuple[int, int, int]:
    """Write calibration and customer documents.

    Returns ``(records, customers, skipped)`` — the last being records an
    engineer deleted from the app, which are passed over rather than rewritten.

    Idempotent: every document id is derived from the file it came from, so a
    re-run updates in place. That is not a nicety — an 87,000-file import will
    be interrupted, and a run that created duplicates on resume would be worse
    than no run at all.

    ``merge=True`` for the same reason: a record the app has since touched keeps
    whatever it added rather than being flattened back to the archive's view.
    """
    from firebase_admin import firestore as fs

    chosen = forms[:limit] if limit else forms

    # ── Customers first ──────────────────────────────────────────────────
    # Records reference a customer, so the site should exist before the
    # calibration that points at it. Categories come from the folder the files
    # were found in; names from the code list.
    # Category from the folder, and the client header from whichever of the
    # site's forms actually carries one — not every device type prints it, so
    # the first form that does wins rather than the first form seen.
    seen: dict[str, str] = {}
    named: dict[str, tuple[str, str]] = {}
    for form in chosen:
        code = form.customer_code
        if not code:
            continue
        if code not in seen:
            seen[code] = form.folder_category or CATEGORY_OTHER
        if code not in named and (form.client_name or form.client_address):
            named[code] = (form.client_name, form.client_address)

    customer_count = 0
    batch = db.batch()
    pending = 0
    for code, category in sorted(seen.items()):
        batch.set(
            db.collection(CUSTOMERS).document(code),
            build_customer(code, codes.get(code), category, named.get(code)),
            merge=True,
        )
        pending += 1
        customer_count += 1
        if pending >= BATCH_SIZE:
            batch.commit()
            batch = db.batch()
            pending = 0
    if pending:
        batch.commit()

    # ── Then the calibrations ────────────────────────────────────────────
    # Counted as *distinct documents*, not files written. Several files can map
    # to one record — copies of the same device's paperwork — and reporting the
    # file count would claim more records than the collection actually holds.
    # Records an engineer deleted from the app. Skipped rather than written, or
    # this run would undo their deletion.
    tombstones = load_tombstones(db)
    if tombstones:
        print(f"  {len(tombstones):,} deleted record(s) will be skipped")

    written: set[str] = set()
    skipped_deleted = 0
    batch = db.batch()
    pending = 0
    for index, form in enumerate(chosen):
        if form.doc_id in tombstones:
            skipped_deleted += 1
            continue

        document = build_document(form, codes.get(form.customer_code))
        document["uploadedAt"] = fs.SERVER_TIMESTAMP
        batch.set(
            db.collection(CALIBRATIONS).document(form.doc_id), document, merge=True
        )
        pending += 1
        written.add(form.doc_id)
        if pending >= BATCH_SIZE:
            batch.commit()
            batch = db.batch()
            pending = 0
            if progress:
                progress(index + 1, len(chosen))
    if pending:
        batch.commit()
    if progress:
        progress(len(chosen), len(chosen))

    return len(written), customer_count, skipped_deleted


def load_tombstones(db) -> set[str]:
    """Record ids an engineer has deleted from the app.

    Read once at the start of a write, because the alternative is checking each
    of 87,000 documents individually. The collection holds one small document
    per deletion and is expected to stay in the hundreds, so pulling it whole
    costs almost nothing.

    Deliberately fails soft: an import that cannot read the tombstones would
    otherwise refuse to run at all, and being unable to honour a deletion is a
    smaller problem than being unable to import. The count is reported so a
    silent zero is never mistaken for "nobody has deleted anything".
    """
    try:
        return {snap.id for snap in db.collection(DELETIONS).stream()}
    except Exception as error:  # noqa: BLE001
        print(f"  WARNING: could not read {DELETIONS}: {type(error).__name__}")
        print("  Deleted records may reappear. Fix this before trusting the run.")
        return set()


def reachable(path: str) -> str:
    """Windows cannot open a path past 260 characters without the ``\\\\?\\`` prefix.

    Six workbooks in the archive nest deeper than that: an automated site files
    its ultrasound probes under folders carrying the full model and serial, so
    ``…\\GE       Logiq P7      LP7352115     Radiology\\Convex      4C     1144687WX9.xlsx``
    runs to 276 characters. ``os.walk`` finds them happily and then ``stat`` and
    ``open`` both raise FileNotFoundError on a file that is plainly there.

    That killed an entire run at the cost-estimate stage — 34 minutes of reading
    thrown away before a single document was written — so the prefix is applied
    here rather than left to each caller to remember.
    """
    if os.name == "nt" and len(path) >= 250 and not path.startswith("\\\\?\\"):
        return "\\\\?\\" + os.path.abspath(path)
    return path


def push_storage(
    forms: list[ParsedForm],
    db,
    storage_module,
    *,
    years: set[int],
    limit: int = 0,
    progress=None,
) -> tuple[int, int]:
    """Upload the original workbooks and record where each one landed.

    Only the years asked for: the decision was metadata for everything, files
    for the recent rounds, with the older years backfillable later by running
    this again with a wider ``--storage-years``.

    Each upload is followed by writing ``storagePath`` back onto the record, so
    a half-finished run leaves the two consistent — a record either has a file
    behind it or does not claim to.
    """
    bucket = storage_module.bucket()
    chosen = [f for f in forms if f.year in years]
    if limit:
        chosen = chosen[:limit]

    # A deleted record must not get its workbook back, and any copy uploaded
    # before the deletion is now an orphan — a certificate for a device that,
    # as far as the engineers are concerned, does not exist. This is where they
    # are cleared: the app cannot do it, because clients have no write access to
    # the bucket and granting it would let any handset erase any certificate.
    tombstones = load_tombstones(db)
    removed = 0

    uploaded = skipped = 0
    for index, form in enumerate(chosen):
        blob_name = f"{STORAGE_PREFIX}/{form.year}/{form.customer_code or 'UNKNOWN'}/{form.filename}"
        blob = bucket.blob(blob_name)

        if form.doc_id in tombstones:
            try:
                if blob.exists():
                    blob.delete()
                    removed += 1
            except Exception as error:  # noqa: BLE001
                print(f"\n  could not clear {blob_name}: {type(error).__name__}")
            continue

        try:
            if blob.exists():
                skipped += 1
            else:
                # Same 900 s as upload_pending: the default 120 is generous
                # for a 129 KB form and hopeless for the handful carrying
                # embedded images, which run to 14 MB.
                blob.upload_from_filename(reachable(form.path), timeout=900)
                uploaded += 1
            db.collection(CALIBRATIONS).document(form.doc_id).set(
                {"storagePath": blob_name}, merge=True
            )
        except Exception as error:  # noqa: BLE001
            debug = f"{type(error).__name__}: {error}"
            print(f"\n  upload failed for {form.filename}: {debug[:90]}")
        if progress and (index % 25 == 0 or index == len(chosen) - 1):
            progress(index + 1, len(chosen))

    return uploaded, skipped, removed


def upload_pending(
    db,
    storage_module,
    *,
    years: set[int],
    jobs: int = 16,
    limit: int = 0,
) -> tuple[int, int, list[tuple[str, str]]]:
    """Finish the Storage upload, working from the records instead of the archive.

    ``push_storage`` above is correct but slow: measured at **74 objects/minute**
    against the real bucket, which is seven hours for one round of the archive.
    At ~129 KB a file that is 159 KB/s — a fraction of the line — so it is not
    bandwidth-bound, it is waiting. Three round trips per file, none overlapping:
    ``exists()``, then the upload, then a single-document ``storagePath`` write.

    This does the same work about fifteen times faster:

    - **The bucket is listed once** into a set, rather than asking the server
      about each file in turn. One listing replaces 47,478 round trips.
    - **``storagePath`` is written in batches of 400.** Firestore takes 500
      operations per round trip, so ~47,000 writes become ~120.
    - **Uploads run on a thread pool.** They are independent and purely
      I/O-bound, so the GIL is released for the whole of each one — the case
      threads are actually good at.

    The work list comes from Firestore rather than a fresh scan, because every
    record already carries ``sourcePath``, ``sourceYear``, ``customerCode`` and
    ``fileName``. Rebuilding it from the archive would mean re-reading 87,000
    workbooks — 35 minutes before a single byte moved.

    Safe to run repeatedly: anything already in the bucket is skipped, and a
    record whose file is up but whose ``storagePath`` never got written (an
    interrupted run) is stamped without re-uploading.

    Returns ``(uploaded, stamped, failures)``.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    import requests.adapters

    bucket = storage_module.bucket()

    # requests defaults to a pool of 10 connections. Workers beyond that queue
    # silently on a free one, so the parallelism is added and then given away.
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=jobs, pool_maxsize=jobs * 2, max_retries=3
    )
    bucket.client._http.mount("https://", adapter)

    print("  listing the bucket…", flush=True)
    present = {blob.name for blob in bucket.list_blobs()}
    print(f"    {len(present):,} objects already there", flush=True)

    print("  reading records…", flush=True)
    pending: list[tuple[str, str, str]] = []      # doc id, blob name, source path
    stamp_only: list[tuple[str, str]] = []        # file is up, record never said so
    for snap in (
        db.collection(CALIBRATIONS)
        .select(["sourceYear", "customerCode", "fileName", "sourcePath", "storagePath"])
        .stream()
    ):
        record = snap.to_dict()
        if record.get("sourceYear") not in years:
            continue
        filename = (record.get("fileName") or "").strip()
        if not filename:
            continue

        site = record.get("customerCode") or "UNKNOWN"
        blob_name = f"{STORAGE_PREFIX}/{record['sourceYear']}/{site}/{filename}"

        if blob_name in present:
            if record.get("storagePath") != blob_name:
                stamp_only.append((snap.id, blob_name))
        else:
            pending.append((snap.id, blob_name, record.get("sourcePath") or ""))

    if limit:
        pending = pending[:limit]

    print(f"    {len(pending):,} to upload, {len(stamp_only):,} to stamp only", flush=True)

    done: list[tuple[str, str]] = []
    failures: list[tuple[str, str]] = []

    def send(job: tuple[str, str, str]) -> tuple[str, str, str | None]:
        doc_id, blob_name, source = job
        try:
            # The SDK defaults to 120 s, which is generous for the ~129 KB the
            # average form weighs and far too short for the handful that carry
            # embedded images — six workbooks run from 5 MB to 14 MB and timed
            # out on every attempt until this was raised.
            bucket.blob(blob_name).upload_from_filename(
                reachable(source), timeout=900
            )
            return doc_id, blob_name, None
        except Exception as error:  # noqa: BLE001
            return doc_id, blob_name, f"{type(error).__name__}: {error}"

    # Timed from here, not from the top: listing the bucket and streaming the
    # records is a fixed couple of minutes, and folding that into the rate makes
    # a fast upload look slow and the estimate useless.
    upload_started = datetime.now()
    if pending:
        with ThreadPoolExecutor(max_workers=jobs) as pool:
            futures = [pool.submit(send, job) for job in pending]
            for index, future in enumerate(as_completed(futures)):
                doc_id, blob_name, error = future.result()
                if error:
                    failures.append((blob_name, error))
                else:
                    done.append((doc_id, blob_name))
                if index % 100 == 0 or index == len(futures) - 1:
                    finished = index + 1
                    mins = max(
                        (datetime.now() - upload_started).total_seconds() / 60, 0.01
                    )
                    rate = finished / mins
                    left = (len(futures) - finished) / rate if rate else 0
                    print(f"\r  {finished:,}/{len(futures):,}  {rate:,.0f}/min  "
                          f"~{left:,.0f} min left    ", end="", flush=True)

    # One round trip per 400 records instead of one per record.
    to_write = done + stamp_only
    if to_write:
        print(f"\n  writing storagePath for {len(to_write):,} records…", flush=True)
        batch = db.batch()
        pending_writes = 0
        for doc_id, blob_name in to_write:
            batch.set(
                db.collection(CALIBRATIONS).document(doc_id),
                {"storagePath": blob_name},
                merge=True,
            )
            pending_writes += 1
            if pending_writes >= 400:
                batch.commit()
                batch = db.batch()
                pending_writes = 0
        if pending_writes:
            batch.commit()

    return len(done), len(stamp_only), failures


# ──────────────────────────────────────────────────────────────────────────────
# The dry-run report
# ──────────────────────────────────────────────────────────────────────────────


@dataclass
class Report:
    """Everything the dry run learned. Rendered as text and as JSON."""

    root: str
    scanned: int = 0
    conforming: int = 0
    known_device: int = 0
    deep_read: int = 0

    #: Split three ways, because they need three different actions.
    #: A code from a properly-named file is a real device worth mapping; the
    #: same-looking string from a file called "ERKA erkameter3000" is just its
    #: first word and mapping it would be meaningless.
    unknown_codes: Counter = field(default_factory=Counter)
    unknown_from_bad_names: Counter = field(default_factory=Counter)
    named_only: Counter = field(default_factory=Counter)
    #: No leading letters to read a code from at all — "2023 report.xlsx" and
    #: the like. Counted so the report reconciles against the scan total; a
    #: summary whose parts do not add up is not one anybody should act on.
    no_code_at_all: int = 0
    device_counts: Counter = field(default_factory=Counter)
    per_year: Counter = field(default_factory=Counter)
    categories: Counter = field(default_factory=Counter)

    serial_kinds: Counter = field(default_factory=Counter)
    suspect_serials: list[tuple[str, str]] = field(default_factory=list)
    duplicate_serials: list[tuple[str, list[str]]] = field(default_factory=list)
    attention: list[tuple[str, str]] = field(default_factory=list)

    codes_in_list: int = 0
    codes_missing_from_list: Counter = field(default_factory=Counter)
    #: code -> an example folder trail, so an unmapped site can be identified.
    missing_code_examples: dict = field(default_factory=dict)
    duplicate_doc_ids: list[tuple[str, list[str]]] = field(default_factory=list)
    skipped_not_forms: int = 0


_DATEISH_RE = re.compile(r"\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}|^\d{4}-\d{2}-\d{2}")

#: Label text -> the field it introduces, for map discovery.
#:
#: Ordered most specific first: "Date of receipt" must beat a bare "Date", and
#: "Control S.N" must not be mistaken for the serial.
_FIELD_LABELS: list[tuple[str, re.Pattern]] = [
    ("S.N", re.compile(r"^serial\s*(no\.?|number)?\s*:?\s*$|^s\.?\s*n\.?\s*:?\s*$", re.I)),
    ("Model", re.compile(r"^model\s*:?\s*$", re.I)),
    ("Manufacturer", re.compile(r"^manufacturer\s*:?\s*$", re.I)),
    ("Location", re.compile(r"^location\s*:?\s*$", re.I)),
    ("Date", re.compile(r"^issue\s*date.*$", re.I)),
    ("Status", re.compile(r"^status\s*:?\s*$", re.I)),
]

_DISCOVER_COLUMNS = "ABCDEFGHIJKLMN"


def discover_map(path: str, max_row: int = 95) -> dict[str, str]:
    """Work out a device's cell map by reading the form's own printed labels.

    Every one of these sheets labels its fields — "Model:", "Serial No.:",
    "Location" — and the value sits in the first cell to the right that holds
    something different. Reading those labels is the only honest way to place a
    new device code: guessing coordinates from a similar form is what produced
    the `AI` map that spent years reporting "ICU" as a serial number.

    Cells are merged heavily, so a label repeats across several columns and so
    does its value. The value is therefore the first column whose text differs
    from the label's, not simply the next column along.
    """
    probe = {
        f"{c}{r}": f"{c}{r}"
        for r in range(1, max_row + 1)
        for c in _DISCOVER_COLUMNS
    }
    values = read_record(path, probe)

    found: dict[str, str] = {}
    for row in range(1, max_row + 1):
        for index, column in enumerate(_DISCOVER_COLUMNS):
            text = values.get(f"{column}{row}", "").strip()
            if not text:
                continue
            for field, pattern in _FIELD_LABELS:
                if field in found or not pattern.match(text):
                    continue
                # Walk right past the rest of the merged label.
                for other in _DISCOVER_COLUMNS[index + 1:]:
                    candidate = values.get(f"{other}{row}", "").strip()
                    if candidate and candidate != text:
                        found[field] = f"{other}{row}"
                        break
                break

    return found


def verify_maps(forms: list[ParsedForm], per_code: int = 6) -> list[str]:
    """Check every device cell map against real files, and report the wrong ones.

    This exists because ``AI`` was wrong. Its map sat two rows too low, so every
    Infusion row in every register Calist has ever produced carried the
    *location* as its serial — "ICU", "NICU" — the manufacturer as its model,
    and nothing at all as its manufacturer. Nobody noticed, because a register
    full of plausible-looking text does not announce that its columns are
    shifted.

    ``device_config`` still carries a TODO listing seven more maps nobody has
    checked. Rather than trust them, this samples real files per code and asks
    three questions that a correct map always answers the same way:

      * does the serial cell contain a digit?
      * does the date cell look like a date?
      * are the model and manufacturer cells non-empty?

    A map that fails the *same* check on every sampled file is misaligned. One
    that fails sporadically is just a form somebody filled in badly, which is a
    different problem and not this function's business.
    """
    by_code: dict[str, list[ParsedForm]] = defaultdict(list)
    for form_ in forms:
        if form_.device_code in DEVICE_CONFIGS:
            by_code[form_.device_code].append(form_)

    findings: list[str] = []

    for code in sorted(by_code):
        config = DEVICE_CONFIGS[code]
        cells = config["cells"]
        sample = by_code[code][:: max(1, len(by_code[code]) // per_code)][:per_code]
        if not sample:
            continue

        read = 0
        bad_serial = bad_date = blank_model = blank_mfr = 0
        rescued = 0

        for form_ in sample:
            try:
                record, layout = read_best(form_.path, config)
            except Exception:  # noqa: BLE001
                continue
            read += 1
            if layout > 0:
                rescued += 1
                # An alternate handled it, so the primary map is not broken —
                # the form simply changed between rounds, which is recorded in
                # device_config and needs no fixing.
                continue
            serial = clean(record.get("S.N"))
            if classify_serial(serial) in ("suspect", "blank"):
                bad_serial += 1
            if not _DATEISH_RE.search(clean(record.get("Date"))):
                bad_date += 1
            if not clean(record.get("Model")):
                blank_model += 1
            if not clean(record.get("Manufacturer")):
                blank_mfr += 1

        # Only the files the primary map actually handled are judged.
        judged = read - rescued
        if judged < 2:
            continue

        # "Every single file the primary map handled" is the bar. Anything less
        # is bad data entry rather than a bad map.
        problems = []
        if bad_serial == judged:
            problems.append(f"serial cell {cells.get('S.N')} never holds a serial")
        if bad_date == judged:
            problems.append(f"date cell {cells.get('Date')} never holds a date")
        if blank_model == judged:
            problems.append(f"model cell {cells.get('Model')} always empty")
        if blank_mfr == judged:
            problems.append(f"manufacturer cell {cells.get('Manufacturer')} always empty")

        if problems:
            findings.append(
                f"{code} ({config['device_name']}, {len(by_code[code]):,} files, "
                f"{judged} judged{f', {rescued} handled by an alternate' if rescued else ''})"
            )
            findings.extend(f"      {p}" for p in problems)

    return findings


# ──────────────────────────────────────────────────────────────────────────────
# Shaping records for Firestore
# ──────────────────────────────────────────────────────────────────────────────

#: The app's own device-type ids, mapped onto the archive's codes.
#:
#: The app shipped with three ids of its own before any of this existed. Rather
#: than rewrite records it has already filed, the archive's vocabulary wins and
#: this records the equivalence, so a balance filed by the app and a BP form
#: from 2023 can be recognised as the same kind of device.
APP_TYPE_TO_CODE = {"balance": "BP", "patient-monitor": "AGH", "x-ray": "BF"}

#: A status cell that means the device passed. Anything else — "Fail", a blank,
#: a comment — is recorded as not passed, because only an explicit pass is one.
_PASS_VALUES = {"pass", "passed", "ok", "accepted", "conform", "conforms"}


def _parse_form_date(text: str) -> str | None:
    """Turn a form's Date cell into an ISO date, or None if it is not one.

    The forms write dates as DD-MM-YYYY. Day-first matters: "05-03-2024" is the
    5th of March, and reading it as the 3rd of May would put the calibration in
    the wrong month and quietly corrupt the previous-calibration lookup this
    whole exercise exists to serve.
    """
    value = text.strip()
    if not value:
        return None
    match = re.match(r"^(\d{1,4})[-/.](\d{1,2})[-/.](\d{1,4})$", value)
    if not match:
        return None
    a, b, c = (int(g) for g in match.groups())
    if a > 31:                      # already year-first
        year, month, day = a, b, c
    else:
        day, month, year = a, b, c
    if year < 100:
        year += 2000
    if not (1 <= month <= 12 and 1 <= day <= 31 and 2000 <= year <= 2099):
        return None
    try:
        return datetime(year, month, day).date().isoformat()
    except ValueError:
        return None


def build_document(form: ParsedForm, customer: dict | None) -> dict:
    """One archive record, in the shape the app's DeviceRecord already reads.

    Field names mirror ``DeviceRecord.toRemoteJson()`` exactly so the app needs
    no second parser — ``DeviceRecord.fromJson`` is defensive on every field, so
    the extra archive-only keys below are simply ignored by it.

    Two identities stay **empty on purpose**: ``engineerUid`` and ``uploadedBy``.
    Nobody knows who performed most of this work, and inventing an attribution
    would both misrepresent the record and — because XP is awarded per synced
    calibration — hand somebody four years of experience they did not earn. An
    empty ``uploadedBy`` also means the security rules refuse any client update,
    which is exactly the read-only behaviour the archive needs.
    """
    issue_iso = _parse_form_date(form.form_date)
    if not issue_iso and form.conforms:
        # Fall back to the filename's MMYY. The day is genuinely unknown, so
        # the 1st is a placeholder — flagged below so nothing reads it as fact.
        issue_iso = f"{2000 + form.filed_yy:04d}-{form.month:02d}-01"
    if not issue_iso:
        issue_iso = f"{form.year:04d}-01-01"

    issue_year = int(issue_iso[:4])
    serial = form.serial.strip()

    attention = form.attention
    if not _parse_form_date(form.form_date) and form.conforms:
        attention = attention or "day of month unknown — taken from the filename"

    document = {
        # ── The shape the app already understands ──
        "id": form.doc_id,
        "serial": serial,
        "customerCode": form.customer_code,
        "deviceType": form.device_code,
        "model": form.model,
        "manufacturer": form.manufacturer,
        "location": form.location,
        "issueDate": issue_iso,
        "issueYear": issue_year,
        "createdAt": issue_iso,
        "updatedAt": issue_iso,
        "passed": form.status.strip().lower() in _PASS_VALUES,
        "fileName": form.filename,
        "engineerCode": "",
        "engineerUid": "",
        "uploadedBy": "",
        # The code list is authoritative for the name where it has one, because
        # it carries the official spelling. It has no addresses at all, and it
        # has never heard of 25 of these sites — so the form fills both gaps.
        "clientName": (customer or {}).get("name", "") or form.client_name,
        "clientAddress": (customer or {}).get("address", "") or form.client_address,
        # Per visit, from the form, which is the right grain: a returning
        # customer gets a fresh contact each year, so this belongs on the
        # calibration rather than on the customer.
        "contactName": form.contact_name,
        "contactPhone": form.contact_phone,
        "tests": {},
        "formData": {},

        # ── Archive-only ──
        "imported": True,
        "sourceYear": form.year,
        "sourcePath": form.path,
        "deviceName": form.device_name,
        "siteTag": form.tag,
        # Site-scoped device key, e.g. "B03-BP002".
        #
        # A tag alone is NOT unique: querying siteTag == "BP002" returns the
        # second balance at every hospital in the country. The previous-
        # calibration lookup has to ask "this device, at this site", and this is
        # the single field that answers it in one indexed query.
        "siteDevice": f"{form.customer_code}-{form.tag}" if form.tag else "",
        # Set only when this device had to be given a new number because
        # several devices shared one filename. The file on disk keeps its
        # original name, so without this a record could not be traced back to
        # the paperwork it came from.
        "originalTag": form.original_tag,
        "renumbered": bool(form.original_tag),
        "deviceCode": form.device_code,
        # The tag's number on its own, e.g. 132 from AH132.
        #
        # Stored as an integer because the app has to find the highest number
        # already used at a site before it can allocate the next one, and
        # string ordering gets that wrong: "BP009" sorts after "BP10". Sorting
        # on this field is the only way to ask the question correctly.
        "tagNumber": form.tag_number,
        # Denormalised for the duplicate guard and the serial lookup, exactly as
        # FirestoreCalibrationStore.upload does for the app's own records.
        "serialUpper": serial.upper(),
        "serialQuality": classify_serial(serial),
        "layoutUsed": form.layout,
        # `storagePath` is deliberately absent, not None.
        #
        # It is written by push_storage once the workbook is actually uploaded,
        # and every write here is a merge. Sending an explicit None would
        # overwrite the real path on every record that already has one —
        # 41,818 of them at the time of writing — and silently detach the whole
        # archive from its files. A field that is simply not mentioned survives
        # the merge untouched.
        "needsAttention": bool(attention),
        "attentionReason": attention,
    }
    return document


def build_customer(
    code: str,
    info: dict | None,
    category: str,
    from_form: tuple[str, str] | None = None,
) -> dict:
    """One customer document, in the shape ``Customer.fromJson`` reads.

    ``contactsByYear`` is left empty: the archive records a contact per *visit*
    inside individual forms, not per customer per year, and inventing one from a
    single form would attach one hospital's contact to every visit it ever had.
    The app fills this in properly the next time an engineer confirms the site.
    """
    return {
        "code": code,
        "category": (info or {}).get("category", category) or CATEGORY_OTHER,
        # The code list wins on the name — it carries the official spelling —
        # but it has no addresses at all and does not know 25 of these sites, so
        # the form's own header fills both gaps.
        "name": (info or {}).get("name", "") or (from_form or ("", ""))[0],
        "address": (info or {}).get("address", "") or (from_form or ("", ""))[1],
        "contactsByYear": {},
        "imported": True,
        # The spreadsheet's own wording, kept so collapsing ten categories into
        # five loses nothing recoverable.
        "sourceCategory": (info or {}).get("sourceCategory", ""),
    }


def _short(path: str, root: Path) -> str:
    """Path relative to the archive root, so a report line fits on a screen."""
    try:
        return str(Path(path).relative_to(root))
    except ValueError:
        return path


def analyse(forms: list[ParsedForm], codes: dict[str, dict], root: Path) -> Report:
    report = Report(root=str(root), scanned=len(forms))

    by_serial: dict[str, list[str]] = defaultdict(list)
    by_doc_id: dict[str, list[str]] = defaultdict(list)

    for form in forms:
        report.per_year[form.year] += 1
        if form.conforms:
            report.conforming += 1
        if form.device_code in DEVICE_CONFIGS:
            report.known_device += 1
            report.device_counts[form.device_name] += 1
        elif named_only(form.device_code):
            report.named_only[form.device_code] += 1
        elif form.device_code:
            if form.conforms:
                report.unknown_codes[form.device_code] += 1
            else:
                report.unknown_from_bad_names[form.device_code] += 1
        else:
            report.no_code_at_all += 1

        if form.folder_category:
            report.categories[form.folder_category] += 1

        if form.attention:
            report.attention.append((form.filename, form.attention))

        if form.customer_code:
            if form.customer_code in codes:
                report.codes_in_list += 1
            else:
                report.codes_missing_from_list[form.customer_code] += 1
                report.missing_code_examples.setdefault(
                    form.customer_code, form.folder_trail
                )

        # The *base* id, before serials disambiguate it. The dry run only
        # samples workbooks, so it cannot know which of these are one device
        # filed twice and which are two devices sharing a filename — it reports
        # the clash, and the import resolves it.
        #
        # Path, not filename: a collision is nearly always the *same* name in
        # two folders, so the name alone tells you nothing about which two
        # files to go and look at.
        by_doc_id[form.base_doc_id].append(_short(form.path, root))

        if form.serial or form.model:  # only meaningful for deep-read forms
            report.deep_read += 1
            kind = classify_serial(form.serial)
            report.serial_kinds[kind] += 1
            if kind == "suspect" and len(report.suspect_serials) < 30:
                report.suspect_serials.append((form.filename, form.serial))
            if kind in ("assigned", "real"):
                # Scoped to the site, not global. An assigned serial like
                # BALANCE001 is only unique within a hospital — the same string
                # at two different sites is two different devices, and flagging
                # that as a collision would bury the real ones.
                by_serial[
                    f"{form.customer_code}|{form.year}|{form.serial.upper()}"
                ].append(form.filename)

    report.duplicate_serials = [
        (key, names) for key, names in by_serial.items() if len(names) > 1
    ][:40]
    report.duplicate_doc_ids = [
        (key, names) for key, names in by_doc_id.items() if len(names) > 1
    ][:40]

    return report


def render(report: Report) -> str:
    """The report as text. Deliberately blunt about what is wrong."""
    out: list[str] = []
    add = out.append

    def rule(title: str = "") -> None:
        add("")
        add(f"── {title} " + "─" * max(0, 66 - len(title)) if title else "─" * 70)

    add(f"MedCal Pro — archive dry run")
    add(f"source: {report.root}")
    add(f"run at: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    rule("Scale")
    add(f"  workbooks scanned          {report.scanned:>8,}")
    for year in sorted(report.per_year):
        add(f"    {year}                     {report.per_year[year]:>8,}")

    rule("Will it parse?")
    total = report.scanned or 1

    def line(label: str, count: int) -> None:
        add(f"  {label:<28} {count:>7,}  ({100 * count / total:4.1f}%)")

    line("filename SITE-TAG-MMYY", report.conforming)
    add("")
    line("readable, device mapped", report.known_device)
    line("identified, not readable", sum(report.named_only.values()))
    line("real code, not mapped yet", sum(report.unknown_codes.values()))
    line("badly named", sum(report.unknown_from_bad_names.values()))
    line("no device code at all", report.no_code_at_all)
    add("  " + "-" * 46)
    accounted = (
        report.known_device
        + sum(report.named_only.values())
        + sum(report.unknown_codes.values())
        + sum(report.unknown_from_bad_names.values())
        + report.no_code_at_all
    )
    line("accounted for", accounted)

    if report.unknown_codes:
        rule("Unmapped device codes — WORTH FIXING")
        add("  These come from correctly-named files, so they are real device")
        add("  codes with a real form behind them. Run --discover on each to")
        add("  read its layout off the printed labels, then add it to")
        add("  device_config.py. Cheaper now than correcting records later.")
        add("")
        for code, count in report.unknown_codes.most_common(20):
            add(f"    {code:<10} {count:>7,}")

    if report.named_only:
        rule("Identified but not readable yet")
        add("  The device type is known and is recorded; the form's layout is")
        add("  not mapped, so no readings are extracted. Imported with")
        add("  needsAttention rather than guessed at.")
        add("")
        for code, count in report.named_only.most_common():
            add(f"    {code:<10} {count:>7,}   {named_only(code)}")

    if report.unknown_from_bad_names:
        rule("Not device codes — badly named files")
        add("  These files do not follow SITE-TAG-MMYY, so what looks like a")
        add("  device code is just the first word of the filename. Renaming the")
        add("  files is the fix; there is nothing to add to the device table.")
        add("")
        for code, count in report.unknown_from_bad_names.most_common(12):
            add(f"    {code:<20} {count:>7,}")

    if report.codes_missing_from_list:
        rule("Site codes not in Hospital Codes.xlsx")
        add("  These sites have calibrations but no entry in the code list, so")
        add("  nothing can name them. The folder trail is shown so you can.")
        add("  Adding them to the spreadsheet before the import is far cheaper")
        add("  than renaming customer documents afterwards.")
        add("")
        for code, count in report.codes_missing_from_list.most_common(25):
            trail = report.missing_code_examples.get(code, "")
            add(f"    {code:<8} {count:>7,}   {trail[:52]}")

    rule("Categories (from folder names)")
    for name, count in report.categories.most_common():
        add(f"    {name:<22} {count:>7,}")

    rule("Top device types")
    for name, count in report.device_counts.most_common(15):
        add(f"    {name:<34} {count:>7,}")

    if report.duplicate_doc_ids:
        rule("Files sharing a SITE-TAG-MMYY name")
        add("  The import tells these apart by serial: same serial means one")
        add("  device filed twice and the copies collapse into one record;")
        add("  different serials mean different devices that happen to share a")
        add("  filename, and each keeps its own record. Nothing is lost either")
        add("  way, and no file on disk is touched.")
        add("")
        for key, names in report.duplicate_doc_ids[:15]:
            add(f"    {key}")
            for n in names[:4]:
                add(f"        {n}")
        add("")
        add("  Same name in two folders is the usual cause — the same device")
        add("  filed under two customers, or a copy left in a working folder.")

    if report.deep_read:
        rule("Serial quality (sampled)")
        add(f"  workbooks opened           {report.deep_read:>8,}")
        add("")
        labels = {
            "real": "manufacturer serial",
            "assigned": "assigned in-house (BALANCE001)",
            "placeholder": "placeholder (N.A, -)",
            "blank": "empty cell",
            "suspect": "NOT a serial (e.g. a location)",
        }
        for kind in ("real", "assigned", "placeholder", "blank", "suspect"):
            count = report.serial_kinds.get(kind, 0)
            pct = 100 * count / report.deep_read
            add(f"    {labels[kind]:<34} {count:>6,}  ({pct:4.1f}%)")

        add("")
        add("  This decides whether K14-by-serial can be trusted. Site+tag")
        add("  identity does not depend on any of it.")

        if report.suspect_serials:
            add("")
            add("  Cells holding something that is not a serial:")
            for name, value in report.suspect_serials[:10]:
                add(f"    {value:<18} {name}")

        if report.duplicate_serials:
            add("")
            add(f"  Same serial twice at the same site ({len(report.duplicate_serials)}):")
            for key, names in report.duplicate_serials[:8]:
                add(f"    {key}")
                for n in names[:3]:
                    add(f"        {n}")

    if report.attention:
        rule(f"Flagged files ({len(report.attention):,} total, first 25)")
        for name, why in report.attention[:25]:
            add(f"    {name:<44} {why}")

    rule()
    add("Nothing was written. Review the above, then re-run with --limit.")
    return "\n".join(out)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────


def _run_push(args, root: Path) -> int:
    """The write path. Deliberately talkative about what it is about to do."""
    codes_path = (
        Path(args.codes) if args.codes else root / "Ashraf" / "Hospital Codes.xlsx"
    )
    codes = load_code_list(codes_path) if codes_path.is_file() else {}
    print(f"code list: {len(codes)} customers")

    print("scanning…")
    forms = scan(root, only_year=args.year, limit=args.limit)
    print(f"  {len(forms):,} forms")

    print("reading workbooks (all of them — this is the slow part)…")

    def tick(done: int, total: int) -> None:
        if done % 200 == 0 or done == total:
            print(f"\r  {done:,}/{total:,}", end="", flush=True)

    deepen(forms, sample=0, progress=tick)
    print()

    # Ids can only be settled once serials are known — that is what tells a
    # duplicate copy apart from two devices sharing a tag.
    collapsed, split = assign_ids(forms)
    distinct = len({f.doc_id for f in forms})
    print(f"  {distinct:,} distinct records from {len(forms):,} files")
    if collapsed:
        print(f"    {collapsed:,} duplicate copies collapsed")
    if split:
        print(f"    {split:,} extra devices recovered from shared filenames")

    if args.year or args.limit:
        print()
        print("  NOTE: renumbering only sees the files in this run. A device")
        print("  number invented here could collide with one from a year that")
        print("  was not scanned. Use a single full run for the real import;")
        print("  --year and --limit are for trials.")

    storage_years = {
        int(y) for y in str(args.storage_years).split(",") if y.strip().isdigit()
    }

    records = len({f.doc_id for f in forms})
    customers = len({f.customer_code for f in forms if f.customer_code})
    files = len([f for f in forms if f.year in storage_years]) if args.storage else 0
    emulator = using_emulator()

    print()
    if emulator:
        print("ABOUT TO WRITE — EMULATOR "
              f"({os.environ['FIRESTORE_EMULATOR_HOST']})")
        print("  Nothing here is billable and nothing reaches the real project.")
    else:
        print(f"ABOUT TO WRITE — LIVE PROJECT '{args.project}'  ** THIS COSTS MONEY **")
    print(f"  calibrations   {records:,} documents into '{CALIBRATIONS}'"
          f"  (from {len(forms):,} files)")
    print(f"  customers      {customers:,} documents into '{CUSTOMERS}'")
    if args.storage:
        print(f"  files          {files:,} workbooks into Storage "
              f"({', '.join(str(y) for y in sorted(storage_years))})")
    else:
        print("  files          none (pass --storage to upload workbooks)")

    print()
    print("  Document ids are derived from the files, so re-running updates")
    print("  rather than duplicating. Existing fields are merged, not replaced.")

    if not emulator:
        # Stated before the prompt, not after, so the number is on screen at
        # the moment the decision is made.
        writes = records + customers
        billable = max(0, writes - 20_000)          # 20k/day free on Blaze
        cost = billable / 100_000 * 0.18
        bytes_ = sum(len(json.dumps(build_document(f, None), default=str))
                     for f in forms[:200]) / min(200, len(forms) or 1)
        stored_mb = records * bytes_ / 1024 / 1024

        print()
        print("  ESTIMATED COST")
        print(f"    writes           {writes:,}  "
              f"({billable:,} billable after the 20k/day free tier)")
        print(f"    one-time         ~${cost:,.2f}")
        print(f"    Firestore data   ~{stored_mb:,.0f} MB "
              f"(indexes add roughly 3x that without fieldOverrides)")
        if files:
            # Guarded as well as prefixed: this is only an estimate, and a file
            # that cannot be measured must never bring down a run that has not
            # yet written anything.
            def _size(form: ParsedForm) -> int:
                try:
                    return Path(reachable(form.path)).stat().st_size
                except OSError:
                    return 0

            gb = sum(_size(f)
                     for f in forms if f.year in storage_years) / 1024**3
            print(f"    Storage          ~{gb:,.2f} GB  "
                  f"→ ~${gb * 0.026:,.2f}/month, ~${files / 10_000 * 0.05:,.2f} to upload")
        print()
        print("  Try it on the emulator first — same code, no cost:")
        print("    firebase emulators:start --only firestore,auth,storage")
        print("    set FIRESTORE_EMULATOR_HOST=localhost:8080")

    print()
    if not args.yes:
        print("Refusing to write without --yes. Re-run with --yes when ready.")
        return 3

    db, storage_module = connect(args.project)

    print("writing records…")
    records, customers, skipped = push_firestore(
        forms, codes, db, limit=args.limit, progress=tick
    )
    print()
    print(f"  {records:,} calibrations, {customers:,} customers")
    if skipped:
        print(f"  {skipped:,} skipped — deleted by an engineer")

    if args.storage:
        print("uploading workbooks…")
        uploaded, skipped, removed = push_storage(
            forms, db, storage_module,
            years=storage_years, limit=args.limit, progress=tick,
        )
        print()
        print(f"  {uploaded:,} uploaded, {skipped:,} already there")
        if removed:
            print(f"  {removed:,} workbook(s) cleared for deleted records")

    print()
    print("Done. Re-run any time — it updates in place.")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="firebase_export",
        description="Export MedCal Pro inspection forms to Firebase.",
    )
    parser.add_argument("root", help='archive root, e.g. "D:/MedCal Pro"')
    parser.add_argument("--dry-run", action="store_true",
                        help="analyse and report; open no connection, write nothing")
    parser.add_argument("--year", type=int, default=None,
                        help="restrict to one year folder")
    parser.add_argument("--limit", type=int, default=0,
                        help="stop after N files (for a trial run)")
    parser.add_argument("--sample", type=int, default=400,
                        help="workbooks to open for serial quality (0 = all)")
    parser.add_argument("--verify-maps", action="store_true",
                        help="check every device cell map against real files")
    parser.add_argument("--discover", metavar="CODE",
                        help="derive a cell map for a device code from its "
                             "forms' printed labels (comma-separate for several)")
    parser.add_argument("--codes", default=None,
                        help="path to Hospital Codes.xlsx")
    parser.add_argument("--report", default=None,
                        help="write the report here (default: alongside this script)")
    parser.add_argument("--push", action="store_true",
                        help="write to Firebase (still needs --yes to proceed)")
    parser.add_argument("--yes", action="store_true",
                        help="confirm the write described by --push")
    parser.add_argument("--project", default="medcalpro",
                        help="Firebase project id (default: medcalpro)")
    parser.add_argument("--storage", action="store_true",
                        help="also upload the workbooks to Cloud Storage")
    parser.add_argument("--storage-years", default="2025,2026",
                        help="which years' workbooks to upload (default: 2025,2026)")
    parser.add_argument("--upload-only", action="store_true",
                        help="finish the Storage upload from the records already in "
                             "Firestore; no archive scan, uploads run in parallel")
    parser.add_argument("--jobs", type=int, default=16,
                        help="parallel uploads for --upload-only (default: 16)")
    args = parser.parse_args(argv)

    # The report carries Arabic folder names and box-drawing rules, and the
    # Windows console still defaults to cp1252, which cannot encode either.
    # Without this the run does all its work and then dies on the last print.
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):  # not a real console; already fine
            pass

    root = Path(args.root)
    if not root.is_dir():
        print(f"error: {root} is not a folder", file=sys.stderr)
        return 2

    if args.upload_only:
        storage_years = {
            int(y) for y in str(args.storage_years).split(",") if y.strip().isdigit()
        }
        print(f"Finishing the Storage upload for {sorted(storage_years)}, "
              f"{args.jobs} at a time.")
        print("Reading the work list from Firestore — the archive is not rescanned.")
        if not args.yes:
            print("\nRefusing to write without --yes. Re-run with --yes when ready.")
            return 3

        db, storage_module = connect(args.project)

        started = datetime.now()
        uploaded, stamped, failures = upload_pending(
            db, storage_module,
            years=storage_years, jobs=args.jobs, limit=args.limit,
        )
        minutes = max((datetime.now() - started).total_seconds() / 60, 0.01)

        print()
        print(f"  {uploaded:,} uploaded, {stamped:,} already up and now recorded")
        print(f"  {minutes:.1f} min wall clock, including the listing and the "
              f"record scan")
        if failures:
            print(f"\n  {len(failures):,} failed:")
            for name, why in failures[:15]:
                print(f"    {name}: {why[:90]}")
            if len(failures) > 15:
                print(f"    … and {len(failures) - 15:,} more")
        print()
        print("Done. Re-run any time — anything already up is skipped.")
        return 0

    if args.discover:
        wanted = {c.strip().upper() for c in args.discover.split(",") if c.strip()}
        print(f"scanning for {', '.join(sorted(wanted))}…")
        forms = scan(root, only_year=args.year)

        for code in sorted(wanted):
            matches = [f for f in forms if f.device_code == code]
            print()
            print(f"── {code} ── {len(matches):,} files")
            if not matches:
                print("   none found")
                continue

            # When the code already has a map, sample the files that map
            # *cannot* read. Sampling broadly finds the layout already
            # configured and tells you nothing — which is exactly what happened
            # with the syringe form, whose failures turned out to sit a row
            # either side of the map in use.
            #
            # Forms genuinely do drift: the same device gets a new template
            # between rounds, and automated exports shift the block again.
            pool = matches
            if code in DEVICE_CONFIGS:
                failing = []
                for form_ in matches[:: max(1, len(matches) // 60)]:
                    try:
                        if read_best(form_.path, DEVICE_CONFIGS[code])[1] < 0:
                            failing.append(form_)
                    except Exception:  # noqa: BLE001
                        continue
                    if len(failing) >= 4:
                        break
                if failing:
                    pool = failing
                    print(f"   sampling {len(failing)} file(s) the current map "
                          f"cannot read")
                else:
                    print("   the current map reads every file sampled")

            # Spread the sample across the archive: the same device filed in
            # 2023 and 2026 may be on two different versions of the form, and a
            # sample taken from one folder would never show it.
            step = max(1, len(pool) // 4)
            agreed: dict[str, Counter] = defaultdict(Counter)
            for form_ in pool[::step][:4]:
                found = discover_map(form_.path)
                print(f"   {Path(form_.filename).name[:34]:<36} "
                      f"{ {k: v for k, v in sorted(found.items())} }")
                for field, ref in found.items():
                    agreed[field][ref] += 1

            print("   ── proposed (most common placement per field) ──")
            proposal = {
                field: counts.most_common(1)[0][0]
                for field, counts in sorted(agreed.items())
            }
            print(f"   {proposal}")
            missing = {f for f, _ in _FIELD_LABELS} - set(proposal)
            if missing:
                print(f"   !! no label found for: {', '.join(sorted(missing))}")
        return 0

    if args.verify_maps:
        print("scanning…")
        forms = scan(root, only_year=args.year, limit=args.limit)
        print(f"  {len(forms):,} forms; checking cell maps against real files…")
        findings = verify_maps(forms)
        print()
        if not findings:
            print("Every device map produces sane values on every sampled file.")
            return 0
        print("MAPS THAT LOOK WRONG")
        print("Each of these failed the same check on every file sampled, which")
        print("means the coordinates are off rather than the data being messy.")
        print()
        for line in findings:
            print(f"  {line}")
        return 1

    if args.push:
        return _run_push(args, root)

    if not args.dry_run:
        print("Nothing to do. Pick a mode:", file=sys.stderr)
        print("  --dry-run       analyse and report, touching nothing", file=sys.stderr)
        print("  --verify-maps   check the device cell maps against real files",
              file=sys.stderr)
        print("  --discover CODE derive a cell map from a form's own labels",
              file=sys.stderr)
        print("  --push          write to Firebase (asks first)", file=sys.stderr)
        return 2

    codes_path = Path(args.codes) if args.codes else root / "Ashraf" / "Hospital Codes.xlsx"
    codes: dict[str, dict] = {}
    if codes_path.is_file():
        try:
            codes = load_code_list(codes_path)
            print(f"code list: {len(codes)} customers from {codes_path.name}")
        except Exception as error:  # noqa: BLE001
            print(f"warning: could not read {codes_path}: {error}", file=sys.stderr)
    else:
        print(f"warning: no code list at {codes_path}", file=sys.stderr)

    print("scanning (no workbooks opened yet)…")
    forms = scan(root, only_year=args.year, limit=args.limit)
    print(f"  {len(forms):,} forms")

    if forms and args.sample != 0:
        wanted = args.sample if args.sample > 0 else 0
        print(f"opening {'all' if not wanted else wanted} workbooks for serial quality…")

        def tick(done: int, total: int) -> None:
            if done % 50 == 0 or done == total:
                print(f"\r  {done:,}/{total:,}", end="", flush=True)

        deepen(forms, sample=wanted, progress=tick)
        print()

    report = analyse(forms, codes, root)
    text = render(report)
    print()
    print(text)

    destination = Path(args.report) if args.report else Path(__file__).with_name(
        f"dry-run-{datetime.now(timezone.utc):%Y%m%d-%H%M}.txt"
    )
    destination.write_text(text, encoding="utf-8")
    destination.with_suffix(".json").write_text(
        json.dumps(
            {
                "root": report.root,
                "scanned": report.scanned,
                "conforming": report.conforming,
                "known_device": report.known_device,
                "deep_read": report.deep_read,
                "serial_kinds": dict(report.serial_kinds),
                "per_year": dict(report.per_year),
                "unknown_codes": dict(report.unknown_codes),
                "named_only": dict(report.named_only),
                "unknown_from_bad_names": dict(report.unknown_from_bad_names),
                "codes_missing_from_list": dict(report.codes_missing_from_list),
                "categories": dict(report.categories),
                "device_counts": dict(report.device_counts),
                "duplicate_doc_ids": report.duplicate_doc_ids,
                "attention_count": len(report.attention),
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    print(f"\nwritten: {destination}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
