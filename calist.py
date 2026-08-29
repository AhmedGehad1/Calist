"""Calist — compile device inspection forms into one equipment register.

Extracts fixed cells from many device-inspection Excel forms and compiles them
into a copy of a pre-defined template.

The device type is read from the filename: everything after the first "-" is
scanned for its leading letters, so "Clinic-AGH001.xlsx" yields the code "AGH".
That code selects a cell map from device_config.py, which says where each field
lives on that device's form. Data is read from the first worksheet that holds
anything — the X-ray workbook opens on an empty macro stub, and reading that
one made every X-ray come out blank.

Output columns (template column B onwards):
    Device | Manufacturer | Model | S.N | Location | Code | Date | Status
"""

from __future__ import annotations

import logging
import os
import re
import sys
import threading
import time
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Iterator
from xml.etree.ElementTree import iterparse

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.numbers import BUILTIN_FORMATS, is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, from_excel

from device_config import DEVICE_CONFIGS

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

#: Fields written to the template, in column order (template column B onwards).
FIELDS = ["Device", "Manufacturer", "Model", "S.N", "Location", "Code", "Date", "Status"]

TEMPLATE_START_ROW = 4          # first data row; rows 1-3 are headers
TEMPLATE_START_COL = 2          # column B (1=A, 2=B, ...)

OUTPUT_NAME = "device list.xlsx"
SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls")

#: A file whose code isn't in DEVICE_CONFIGS is skipped with an error. Set this
#: to False to fall back to UNKNOWN_FALLBACK instead (the old behaviour), which
#: emits a row of whatever happens to be in A1:A6.
SKIP_UNKNOWN_CODES = True

UNKNOWN_FALLBACK = {
    "device_name": "Unknown Device",
    "cells": {
        "Manufacturer": "A1", "Model": "A2", "S.N": "A3",
        "Location": "A4", "Date": "A5", "Status": "A6",
    },
}

#: Device pairs permitted to share one serial number, because the second row is
#: generated from the same physical unit. Names must match device_config.py.
ALLOWED_SHARED_SN_PAIRS = {
    frozenset({"Patient Monitor", "NIBP"}),
    frozenset({"Vital Sign (SPO2 Module)", "Vital Sign (NIBP Module)"}),
}

Record = dict[str, str]

log = logging.getLogger("aggregator")

#: Where the reference template lives relative to the app.
TEMPLATE_NAME = "Device List.xlsx"

#: The single source of the version number. calist.spec reads it straight out
#: of this file to stamp the executable's Windows version resource, so the
#: About box and the file's Properties tab cannot drift apart.
__version__ = "1.6.0"

#: Authorship. Written into every register and into the workbook's document
#: properties, so the credit travels with the file rather than living only in
#: the app that made it.
AUTHOR_NAME = "Ahmed Gehad"
AUTHOR_EMAIL = "ahmedgehad2112@gmail.com"
ATTRIBUTION = f"{AUTHOR_NAME} · {AUTHOR_EMAIL}"


def bundled_template() -> Path | None:
    """The reference template shipped with the app, if it can be found.

    A frozen build unpacks its data files to ``sys._MEIPASS``; running from a
    checkout, they sit beside this module. Returning it as a default means a
    freshly downloaded copy is usable without hunting for a template first.
    """
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidate = base / "template" / TEMPLATE_NAME
    return candidate if candidate.is_file() else None


# ──────────────────────────────────────────────────────────────────────────────
# Structured results
#
# The pipeline reports progress twice over: as log records (the seam the GUI
# and headless callers both read), and as these dataclasses, which carry the
# same facts in a form a table can render. Neither replaces the other.
# ──────────────────────────────────────────────────────────────────────────────

#: A file's state. "ready" is pre-flight only — it means the filename resolved
#: to a known device, not that the workbook has been opened.
READY = "ready"
OK = "ok"
UNSUPPORTED = "unsupported"
UNKNOWN_CODE = "unknown_code"
BAD_FORMAT = "bad_format"
ERROR = "error"
CANCELLED = "cancelled"

#: Statuses that mean a file contributed nothing to the register.
PROBLEM_STATUSES = frozenset({UNSUPPORTED, UNKNOWN_CODE, BAD_FORMAT, ERROR})


@dataclass
class FileOutcome:
    """What happened to one source file."""

    filename: str
    path: str
    status: str
    device_code: str | None = None
    device_name: str | None = None
    #: The serial as read from the form. Empty until the file is actually
    #: opened — pre-flight resolves a device from the filename alone and never
    #: touches the workbook, so it cannot know this.
    serial: str = ""
    detail: str = ""
    rows: int = 0               # 1, or 2 where a second_row was generated

    @property
    def is_problem(self) -> bool:
        return self.status in PROBLEM_STATUSES


@dataclass
class Duplicate:
    """A record dropped because its serial number was already recorded.

    The two filenames are the point: the user has to open both to work out
    which one is wrong, and the device type cannot tell them apart when a round
    holds a dozen of the same model.
    """

    serial: str
    device: str
    dropped: str                # the file whose row was dropped
    kept: str                   # the file that recorded the serial first


@dataclass
class RunResult:
    """The outcome of one full run."""

    output_path: Path | None = None
    rows_written: int = 0
    outcomes: list[FileOutcome] = field(default_factory=list)
    duplicates_removed: int = 0
    #: The dropped records themselves. The log carries the same facts; a
    #: summary needs them in a form it can render without scraping text.
    duplicates: list[Duplicate] = field(default_factory=list)
    second_rows_added: int = 0
    cancelled: bool = False
    error: str | None = None

    @property
    def succeeded(self) -> bool:
        return self.output_path is not None

    @property
    def files_read(self) -> int:
        return sum(1 for o in self.outcomes if o.status == OK)

    @property
    def problems(self) -> list[FileOutcome]:
        return [o for o in self.outcomes if o.is_problem]


#: Called after each file with (outcome, index, total). Fires on the worker
#: thread, so a GUI caller must marshal back to its own loop.
ProgressHook = Callable[[FileOutcome, int, int], None]


# ──────────────────────────────────────────────────────────────────────────────
# Value normalisation
# ──────────────────────────────────────────────────────────────────────────────

def clean(value: object) -> str:
    """Render a raw cell value as the string that belongs in the output.

    Straight str() of whatever Excel hands back. Note this means a real date
    cell is written as "2024-01-15 00:00:00" and a numeric serial as
    "123456.0"; the forms in use hold these as text, so it does not arise.
    """
    return "" if value is None else str(value).strip()


# ──────────────────────────────────────────────────────────────────────────────
# Filename → device code
# ──────────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────────────────
# Optional filename-format check
#
#   G302  -  AGH001  -  0425
#   │        │          └── MMYY: month 01-12, then a two-digit year
#   │        └───────────── device code + unit number
#   └────────────────────── site code: letters then digits
#
# The happy path is ONE precompiled match and nothing else — no splitting, no
# allocation — because this runs on every file the moment it is added. The
# per-part diagnosis below only runs for names that already failed, which costs
# nothing on a folder where everything is named correctly.
# ──────────────────────────────────────────────────────────────────────────────

FILENAME_EXAMPLE = "G302-AGH001-0425"

_NAME_RE = re.compile(r"[A-Za-z]+\d+-[A-Za-z]+\d+-(?:0[1-9]|1[0-2])\d{2}\Z")
_PART_RE = re.compile(r"[A-Za-z]+\d+\Z")
_DATE_RE = re.compile(r"\d{4}\Z")


def check_filename_format(stem: str) -> str | None:
    """None if the stem matches the house format, else why it doesn't."""
    if _NAME_RE.match(stem):
        return None
    return _explain_bad_filename(stem)


def _explain_bad_filename(stem: str) -> str:
    """Say which part is wrong. Only reached for names that already failed."""
    parts = stem.split("-")
    if len(parts) != 3:
        return (f"expected 3 parts like {FILENAME_EXAMPLE}, "
                f"found {len(parts)}")

    site, device, date = parts
    if not _PART_RE.match(site):
        return f"site code '{site}' should be letters then digits, like G302"
    if not _PART_RE.match(device):
        return f"device code '{device}' should be letters then digits, like AGH001"
    if not _DATE_RE.match(date):
        return f"date '{date}' should be 4 digits (MMYY), like 0425"
    if not 1 <= int(date[:2]) <= 12:
        return f"month '{date[:2]}' in '{date}' is not between 01 and 12"
    return f"does not match {FILENAME_EXAMPLE}"


def classify_file(filepath: str, strict_names: bool = False) -> FileOutcome:
    """Work out what a file *would* produce, without opening it.

    Extension check, optional filename-format check, and a filename-to-config
    lookup — no workbook I/O, so this is cheap enough to run on every file the
    moment it is added. That is what lets a bad name or an unrecognised device
    surface before a long run rather than after it.

    With ``strict_names`` the house format is enforced first: when it is on,
    the user has asked for that shape specifically, so a malformed name is the
    finding worth reporting even if a device code could still be salvaged.
    """
    path = Path(filepath)
    filename = path.name

    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        return FileOutcome(filename, filepath, UNSUPPORTED,
                           detail=f"Unsupported format '{path.suffix}'")

    if strict_names:
        problem = check_filename_format(path.stem)
        if problem:
            return FileOutcome(filename, filepath, BAD_FORMAT, detail=problem)

    code = extract_device_code(filename)
    config = DEVICE_CONFIGS.get(code)

    if config is None:
        shown = f"'{code}'" if code else "none found"
        return FileOutcome(filename, filepath, UNKNOWN_CODE, device_code=code,
                           detail=f"code {shown} is not in the device table")

    name = config["device_name"]
    rows = 2 if "second_row" in config else 1
    if rows == 2:
        name = f"{name}  +{config['second_row']['device_name']}"

    return FileOutcome(filename, filepath, READY, device_code=code,
                       device_name=name, rows=rows)


def extract_device_code(filename: str) -> str | None:
    """Pull the device-type code out of a filename.

    Everything after the first "-" is used when one is present, otherwise the
    whole stem; the leading run of letters is the code.

        "Clinic-AGH001.xlsx" -> "AGH"
        "VNT023.xlsx"        -> "VNT"
    """
    stem = Path(filename).stem
    _, separator, tail = stem.partition("-")
    match = re.match(r"[A-Za-z]+", tail if separator else stem)
    return match.group(0).upper() if match else None


# ──────────────────────────────────────────────────────────────────────────────
# Finding forms in a folder
# ──────────────────────────────────────────────────────────────────────────────

#: Excel drops a lock file beside every workbook someone has open, named "~$"
#: plus the workbook's name. They are not workbooks — opening one fails with
#: "File is not a zip file" — so a folder anyone is working in used to produce
#: a row of failures for files the user never chose.
LOCK_PREFIX = "~$"

#: Windows FILE_ATTRIBUTE_HIDDEN. os.scandir fills st_file_attributes straight
#: from the directory listing on Windows, so testing it costs no extra call.
_HIDDEN = 0x2


def is_source_file(name: str) -> bool:
    """Whether a filename found by scanning a folder is worth opening.

    Deliberately not applied to files the user picked by hand: choosing a .docx
    should still report "Unsupported format", because saying nothing at all
    about a file someone explicitly selected is worse than an error row.
    """
    lowered = name.lower()
    return (not name.startswith(LOCK_PREFIX)
            and lowered != OUTPUT_NAME.lower()
            and lowered.endswith(SUPPORTED_EXTENSIONS))


def find_source_files(
    folder: str | os.PathLike,
    cancel: threading.Event | None = None,
) -> Iterator[str]:
    """Yield every form under ``folder``, depth first.

    os.scandir hands back entries that already know file-from-directory and,
    on Windows, their attributes too — so the extension test runs before any
    stat and nothing is materialised in full. ``Path.rglob("*")`` followed by
    ``is_file()`` costs a stat per entry instead: tolerable on a local disk,
    minutes on a network share, and all of it on whichever thread called it.

    Unreadable sub-folders are skipped rather than raising, so one permission
    error in a deep tree cannot lose the rest of the round.
    """
    stack = [os.fspath(folder)]
    while stack:
        if cancel is not None and cancel.is_set():
            return
        try:
            entries = os.scandir(stack.pop())
        except OSError as exc:
            log.debug("Skipped a folder while scanning: %s", exc)
            continue

        with entries:
            while True:
                try:
                    entry = next(entries)
                except StopIteration:
                    break
                except OSError as exc:              # pragma: no cover
                    log.debug("Stopped reading a folder: %s", exc)
                    break

                try:
                    if entry.is_dir(follow_symlinks=False):
                        stack.append(entry.path)
                        continue
                    if not is_source_file(entry.name):
                        continue
                    if getattr(entry.stat(), "st_file_attributes", 0) & _HIDDEN:
                        continue
                except OSError:                     # vanished mid-scan
                    continue
                yield entry.path


# ──────────────────────────────────────────────────────────────────────────────
# Reading source workbooks
# ──────────────────────────────────────────────────────────────────────────────

CellGetter = Callable[[str], object]

# ── The .xlsx/.xlsm package, read directly ────────────────────────────────────
#
# load_workbook() costs ~500ms on these forms — it parses every worksheet, the
# whole styles table, the drawings and the calc chain, to reach seven cells.
# Reading the package directly costs ~2ms: the workbook relationships, one
# sheet's bytes, and the shared strings or the styles only when a wanted cell
# turns out to need them.
#
# Everything below reproduces openpyxl's own answers deliberately — which sheet
# it would have picked, how it renders a number or a date, how it looks through
# a merged range. A whole-grid comparison against the old reader is the check
# that keeps it honest; see the reader tests in test_calist.py.

_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_DOC_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"

_WORKBOOK_PART = "xl/workbook.xml"
_WORKBOOK_RELS = "xl/_rels/workbook.xml.rels"
_SHARED_STRINGS_PART = "xl/sharedStrings.xml"
_STYLES_PART = "xl/styles.xml"

#: One cell element, matched from the offset of its own opening tag.
#:
#: The lazy quantifier on the attribute run is load-bearing. Greedy, it eats
#: the "/" of a self-closing ``<c r="E18" s="168"/>``, then matches the ">"
#: branch and swallows everything up to the *next* cell's ``</c>`` — silently
#: returning a neighbouring cell's value. test_a_self_closing_cell_does_not_
#: swallow_the_next_one pins this.
_CELL_RE = re.compile(rb'<c r="[A-Z]+\d+"([^>]*?)(?:/>|>(.*?)</c>)', re.S)
_CELL_TYPE_RE = re.compile(rb'\bt="([^"]+)"')
_CELL_STYLE_RE = re.compile(rb'\bs="(\d+)"')
_VALUE_RE = re.compile(rb'<v[^>]*>(.*?)</v>', re.S)
_TEXT_RUN_RE = re.compile(rb'<t[^>]*>(.*?)</t>', re.S)
_MERGE_RE = re.compile(rb'<mergeCell\s+ref="([^"]+)"')
_OUTSIDE_A1_RE = re.compile(rb'<c r="(?!A1")')

_ENTITY_RE = re.compile(rb'&(?:#(\d+)|#x([0-9a-fA-F]+)|(amp|lt|gt|quot|apos));')
_NAMED_ENTITIES = {b"amp": b"&", b"lt": b"<", b"gt": b">",
                   b"quot": b'"', b"apos": b"'"}


def _unescape(raw: bytes) -> str:
    """Decode the entities an XML parser would have decoded for us."""
    if b"&" not in raw:
        return raw.decode("utf-8")

    def replace(match: re.Match) -> bytes:
        decimal, hexadecimal, named = match.groups()
        if decimal:
            return chr(int(decimal)).encode("utf-8")
        if hexadecimal:
            return chr(int(hexadecimal, 16)).encode("utf-8")
        return _NAMED_ENTITIES[named]

    return _ENTITY_RE.sub(replace, raw).decode("utf-8")


def _cast_number(text: str) -> int | float:
    """openpyxl's rule for a numeric cell, reproduced.

    Kept here rather than imported from openpyxl's private worksheet reader so
    that a version bump cannot quietly change how a number reaches clean() —
    "123456" must stay an int and "123456.0" a float, because that difference
    is what ends up written into the register.
    """
    if "." in text or "E" in text or "e" in text:
        return float(text)
    return int(text)


def _sheet_has_content(data: bytes) -> bool:
    """True unless this worksheet is unambiguously empty.

    The openpyxl equivalent asked whether max_row/max_column reach past A1, or
    A1 itself holds a value. Both count a *styled* but valueless cell, so this
    looks for cell elements, not for values — a sheet carrying only formatting
    is kept, exactly as before.
    """
    if _OUTSIDE_A1_RE.search(data):
        return True
    at = data.find(b'<c r="A1"')
    if at < 0:
        return False
    match = _CELL_RE.match(data, at)
    body = match.group(2) if match else None
    return bool(body) and (b"<v" in body or b"<is" in body)


class _XlsxSource:
    """One worksheet of an .xlsx/.xlsm package, held as raw XML."""

    def __init__(self, filepath: str):
        self._archive = zipfile.ZipFile(filepath)
        try:
            self._parts, self._epoch = self._workbook_parts()
            self._data = self._populated_sheet(self._parts)
            self._check_attribute_order()
            self._merges = self._merged_ranges()
            self._strings: list[str] | None = None
            self._date_styles: dict[int, bool] = {}
        except Exception:
            self._archive.close()
            raise

    def _check_attribute_order(self) -> None:
        """Excel, LibreOffice and openpyxl all write r as a cell's first
        attribute, and the lookups below rely on it. Two memchr scans
        (~0.02 ms) prove it for the file in hand rather than assuming it,
        because the failure would otherwise be a silently blank field.
        """
        if self._data.count(b"<c ") != self._data.count(b'<c r="'):
            raise ValueError(
                "cell references are not written where this reader can "
                "find them (unexpected attribute order)")

    def select_sheet(self, name: str) -> bool:
        """Re-point this source at another tab of the same workbook.

        Used only when the sheet chosen on open turns out not to carry the
        device details — GC files put them on 'Data entry' behind a 'Report'
        tab, BM files on 'cover page'. Re-reading one part of an already-open
        archive is far cheaper than opening the file again.
        """
        for candidate, part in self._parts:
            if candidate != name:
                continue
            self._data = self._archive.read(part)
            self._check_attribute_order()
            self._merges = self._merged_ranges()
            self.sheet_name = name
            return True
        return False

    def close(self) -> None:
        self._archive.close()

    # ── package structure ────────────────────────────────────────────────────

    def _workbook_parts(self) -> tuple[list[tuple[str, str]], datetime]:
        """The worksheet names and parts, in the order openpyxl would list them.

        Two rules matter and are easy to get wrong. An older .xlsm carries its
        VBA modules as <sheet> entries with an empty r:id — openpyxl drops
        those with a warning, so they must not shift the ordering here.
        Chartsheets go to wb.chartsheets rather than wb.worksheets and are
        skipped; **dialogsheets are not** — openpyxl counts them, and the X-ray
        workbook opens on one.
        """
        targets: dict[str, tuple[str, str]] = {}
        with self._archive.open(_WORKBOOK_RELS) as handle:
            for _, element in iterparse(handle):
                if element.tag == _PKG_REL_NS + "Relationship":
                    kind = (element.get("Type") or "").rsplit("/", 1)[-1]
                    targets[element.get("Id")] = (kind, element.get("Target"))

        parts: list[tuple[str, str]] = []
        epoch = WINDOWS_EPOCH
        with self._archive.open(_WORKBOOK_PART) as handle:
            for _, element in iterparse(handle):
                if element.tag == _MAIN_NS + "workbookPr":
                    if element.get("date1904") in ("1", "true"):
                        epoch = MAC_EPOCH
                elif element.tag == _MAIN_NS + "sheet":
                    rel = element.get(_DOC_REL_NS + "id")
                    if not rel:
                        continue
                    kind, target = targets.get(rel, (None, None))
                    if target is None or kind == "chartsheet":
                        continue
                    parts.append((element.get("name") or "", _resolve_part(target)))

        if not parts:
            raise ValueError("workbook has no worksheets")
        return parts, epoch

    def _populated_sheet(self, parts: list[tuple[str, str]]) -> bytes:
        """The first worksheet that holds anything at all.

        Always reading sheet one is right for almost every form. The X-ray
        workbook opens on an empty ``Waveform Dialog`` stub left behind by its
        macros, with the real form on the next tab — so every mapped cell read
        blank, on every X-ray, with no error anywhere.

        The names are kept for --inspect, which has to be able to say which tab
        it actually read; a diagnostic that names the wrong sheet is worse than
        none.
        """
        self.sheet_names = [name for name, _ in parts]
        self.sheet_name = self.sheet_names[0]
        self.skipped_sheets: list[str] = []

        first: bytes | None = None
        for name, part in parts:
            data = self._archive.read(part)
            if first is None:
                first = data
            if _sheet_has_content(data):
                self.sheet_name = name
                return data
            self.skipped_sheets.append(name)

        self.skipped_sheets = []
        return first if first is not None else b""

    def merged_refs(self) -> list[str]:
        """Every merged range on the sheet that was read, as A1 references."""
        return sorted(
            f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"
            for top, left, bottom, right in self._merges)

    def _merged_ranges(self) -> list[tuple[int, int, int, int]]:
        ranges = []
        for ref in _MERGE_RE.findall(self._data):
            start, _, end = ref.decode().partition(":")
            top, left = coordinate_to_tuple(start)
            bottom, right = coordinate_to_tuple(end or start)
            ranges.append((top, left, bottom, right))
        return ranges

    # ── cells ────────────────────────────────────────────────────────────────

    def _anchor(self, ref: str) -> str:
        """Look a reference through to the top-left of any range covering it.

        The forms draw each answer as a box spanning two columns, and a merged
        range stores its value only in that top-left cell. A cell map naming
        the second column of a box (``L17`` of a merged ``K17:L17``) otherwise
        reads blank with no error anywhere — which is how the Ultrasound serial
        number disappeared when that form was re-laid-out.
        """
        row, column = coordinate_to_tuple(ref)
        for top, left, bottom, right in self._merges:
            if top <= row <= bottom and left <= column <= right:
                if row == top and column == left:
                    return ref
                return f"{get_column_letter(left)}{top}"
        return ref

    def values(self, refs: Iterable[str]) -> dict[str, object]:
        """Read the given references, resolving merges and shared strings."""
        anchors = {ref: self._anchor(ref) for ref in refs}

        raw: dict[str, tuple[str, int, bytes | None]] = {}
        for anchor in set(anchors.values()):
            at = self._data.find(b'<c r="' + anchor.encode() + b'"')
            if at < 0:
                continue
            match = _CELL_RE.match(self._data, at)
            if match is None:
                continue
            attributes = match.group(1) or b""
            kind = _CELL_TYPE_RE.search(attributes)
            style = _CELL_STYLE_RE.search(attributes)
            raw[anchor] = (kind.group(1).decode() if kind else "n",
                           int(style.group(1)) if style else 0,
                           match.group(2))

        return {ref: self._value(*raw[anchor]) if anchor in raw else None
                for ref, anchor in anchors.items()}

    def _value(self, kind: str, style_id: int, body: bytes | None) -> object:
        """Render one cell the way openpyxl's data_only reader would."""
        if not body:
            return None

        if kind == "inlineStr":
            if b"<rPh" in body:
                # Phonetic runs; openpyxl drops them and keeps the base text.
                # Rather than guess, say so — a wrong field is worse than a
                # named failure.
                raise ValueError("cell carries phonetic runs")
            return "".join(_unescape(run) for run in _TEXT_RUN_RE.findall(body))

        # data_only semantics: a formula cell yields its *cached* result, so a
        # file written by a script and never opened in Excel has none.
        match = _VALUE_RE.search(body)
        if match is None:
            return None
        text = match.group(1)

        if kind == "s":
            return self._shared_string(int(text))
        if kind == "b":
            return bool(int(text))
        if kind in ("str", "e", "d"):
            return _unescape(text)

        number = _cast_number(text.decode())
        if style_id and self._is_date_style(style_id):
            try:
                return from_excel(number, self._epoch)
            except (OverflowError, ValueError):
                return "#VALUE!"
        return number

    # ── side parts, read only when a wanted cell needs them ──────────────────

    def _shared_string(self, index: int) -> str:
        if self._strings is None:
            self._strings = self._read_shared_strings()
        try:
            return self._strings[index]
        except IndexError:
            return ""

    def _read_shared_strings(self) -> list[str]:
        try:
            handle = self._archive.open(_SHARED_STRINGS_PART)
        except KeyError:
            return []
        strings: list[str] = []
        runs: list[str] = []
        with handle:
            for _, element in iterparse(handle, ("end",)):
                if element.tag == _MAIN_NS + "t":
                    runs.append(element.text or "")
                elif element.tag == _MAIN_NS + "si":
                    strings.append("".join(runs))
                    runs.clear()
                    element.clear()
        return strings

    def _is_date_style(self, style_id: int) -> bool:
        """Whether this style's number format makes the cell a date.

        Only reached for a numeric cell that a cell map actually asks for, so
        the styles table — 184KB on these forms — is usually never read at all.
        """
        if style_id not in self._date_styles:
            self._date_styles = self._read_date_styles()
        return self._date_styles.get(style_id, False)

    def _read_date_styles(self) -> dict[int, bool]:
        try:
            handle = self._archive.open(_STYLES_PART)
        except KeyError:
            # No styles table at all: nothing can be a date, and a missing
            # part must not take the whole file down.
            return {}

        formats = dict(BUILTIN_FORMATS)
        applied: list[int] = []
        inside = False
        with handle:
            for event, element in iterparse(handle, ("start", "end")):
                if event == "start":
                    if element.tag == _MAIN_NS + "cellXfs":
                        inside = True
                    continue
                if element.tag == _MAIN_NS + "numFmt":
                    formats[int(element.get("numFmtId"))] = element.get("formatCode")
                elif element.tag == _MAIN_NS + "cellXfs":
                    break          # dxfs follows, and can be far larger
                elif element.tag == _MAIN_NS + "xf" and inside:
                    applied.append(int(element.get("numFmtId", 0)))
                element.clear()
        return {index: is_date_format(formats.get(number) or "")
                for index, number in enumerate(applied)}


def _resolve_part(target: str) -> str:
    """A relationship target as a package path."""
    if target.startswith("/"):
        return target[1:]
    return target if target.startswith("xl/") else "xl/" + target


# ── .xls, through xlrd ────────────────────────────────────────────────────────

class _XlsSource:
    """One sheet of a legacy .xls, read through xlrd."""

    def __init__(self, filepath: str):
        # formatting_info is what carries the merge list. It costs more memory
        # and some files refuse it, which must not turn into a read failure —
        # without it merged_cells is simply empty and behaviour is as before.
        try:
            self._workbook = xlrd.open_workbook(filepath, formatting_info=True,
                                                on_demand=True)
        except Exception:
            log.debug("No formatting info for %s; merges unresolved",
                      filepath, exc_info=True)
            self._workbook = xlrd.open_workbook(filepath, on_demand=True)
        self._sheet = self._populated_sheet()
        self._merges = getattr(self._sheet, "merged_cells", ()) or ()
        self.sheet_names = self._workbook.sheet_names()
        self.sheet_name = self._sheet.name

    def select_sheet(self, name: str) -> bool:
        """Re-point this source at another sheet — see _XlsxSource.select_sheet."""
        for index in range(self._workbook.nsheets):
            if self._workbook.sheet_names()[index] != name:
                continue
            self._sheet = self._workbook.sheet_by_index(index)
            self._merges = getattr(self._sheet, "merged_cells", ()) or ()
            self.sheet_name = name
            return True
        return False

    def close(self) -> None:
        try:
            self._workbook.release_resources()
        except Exception:                                # pragma: no cover
            pass

    def _populated_sheet(self):
        """The first sheet holding anything, loading one at a time.

        workbook.sheets() would load every sheet up front, which is the whole
        cost on_demand=True exists to avoid.
        """
        first = None
        for index in range(self._workbook.nsheets):
            sheet = self._workbook.sheet_by_index(index)
            if first is None:
                first = sheet
            if sheet.nrows and sheet.ncols:
                return sheet
            if sheet is not first:
                self._workbook.unload_sheet(index)
        if first is None:
            raise ValueError("workbook has no sheets")
        return first

    def values(self, refs: Iterable[str]) -> dict[str, object]:
        return {ref: self._value(ref) for ref in refs}

    def _value(self, ref: str) -> object:
        row, column = coordinate_to_tuple(ref)
        value = self._at(row, column)
        if value not in (None, ""):
            return value
        # xlrd ranges are 0-based with exclusive upper bounds.
        for top, bottom, left, right in self._merges:
            if top <= row - 1 < bottom and left <= column - 1 < right:
                return self._at(top + 1, left + 1)
        return value

    def _at(self, row: int, column: int) -> object:
        try:
            # Raw value: xlrd reports a date as a bare Excel serial number,
            # which clean() renders as e.g. "45306.0".
            return self._sheet.cell_value(row - 1, column - 1)
        except IndexError:
            return None


def _open_workbook(filepath: str):
    if Path(filepath).suffix.lower() == ".xls":
        return _XlsSource(filepath)
    return _XlsxSource(filepath)


@contextmanager
def _open_source(filepath: str) -> Iterator[CellGetter]:
    """Yield a function that reads a cell (by A1 reference) from one workbook.

    A compatibility seam, not the hot path: read_record asks for the whole cell
    map in one call, because resolving merges once for the batch is what makes
    a form cost two milliseconds instead of five hundred.
    """
    source = _open_workbook(filepath)
    try:
        cache: dict[str, object] = {}

        def get(ref: str) -> object:
            if ref not in cache:
                cache.update(source.values([ref]))
            return cache.get(ref)

        yield get
    finally:
        source.close()


def read_record(filepath: str, cell_map: dict[str, str]) -> Record:
    """Read every mapped cell out of one source file."""
    source = _open_workbook(filepath)
    try:
        values = source.values({ref for ref in cell_map.values() if ref})
    finally:
        source.close()
    return {
        field: clean(values.get(ref)) if ref else ""
        for field, ref in cell_map.items()
    }


# ──────────────────────────────────────────────────────────────────────────────
# Reading a form whose map no longer fits
#
# Measured over the 2025 and 2026 rounds — 42,826 forms — the configured map
# works for 94.7% of them. The rest are not broken files: they are forms that
# were re-laid-out between visits, and the shift is uniform, usually one to
# three rows. `alt_cells` records the layouts we know; this is the net for the
# ones nobody has written down yet.
# ──────────────────────────────────────────────────────────────────────────────

#: A serial the site assigned because the device carries none, e.g. BALANCE001.
_ASSIGNED_SERIAL_RE = re.compile(r"^[A-Za-z][A-Za-z ]*\.?\s*\d+$")

#: The test is *any* digit, not a trailing one: STX21170332PA and NV03124H are
#: perfectly good manufacturer serials that end in a letter. What separates a
#: serial from a location typed into the wrong cell is having numbers at all.
_HAS_DIGIT_RE = re.compile(r"\d")

#: Written where a device genuinely has no serial and none was assigned.
#:
#: "0.0" and "0.00" are here because a numeric zero cell arrives as a float and
#: renders that way — a Baby Warmer cover sheet filled in with zeros throughout
#: otherwise reads as a real serial and gets imported as a device.
_PLACEHOLDERS = {"", "-", "--", "n.a", "na", "n/a", "none", "null",
                 "0", "0.0", "0.00", "00"}

#: A date, in any shape these forms use. Excluded by name because a misaligned
#: map lands on a date more often than on anything else, and "16-01-2026" would
#: otherwise pass a digits-only test and be imported as a serial number.
_DATE_LIKE_RE = re.compile(r"^\s*\d{1,4}\s*[-/.]\s*\d{1,2}\s*[-/.]\s*\d{1,4}\s*$")


def classify_serial(serial: str) -> str:
    """One of: blank, placeholder, assigned, real, suspect.

    ``suspect`` is the finding that matters: the cell holds something that is
    not a serial by any reading — the archive has infusion pumps whose serial
    cell says "ICU" — which is how a misaligned map announces itself.
    """
    value = serial.strip()
    if not value:
        return "blank"
    if value.lower() in _PLACEHOLDERS:
        return "placeholder"
    if _DATE_LIKE_RE.match(value):
        return "suspect"
    if not _HAS_DIGIT_RE.search(value):
        return "suspect"
    if _ASSIGNED_SERIAL_RE.match(value):
        return "assigned"
    return "real"


def plausible(record: Record) -> bool:
    """A record is plausible when the serial looks like one and a model is set.

    Two fields rather than one: a misaligned map that happens to land on some
    other number would pass a serial-only check, and one that lands on a label
    would pass a model-only check. Requiring both is what separates layouts.
    """
    if not record:
        return False
    if classify_serial(clean(record.get("S.N"))) in ("suspect", "blank"):
        return False
    model = clean(record.get("Model"))
    if not model:
        return False
    # A record that is placeholders all the way through is not a reading of
    # anything. Some cover sheets are filled in with "0" throughout, and a
    # serial of "0" classifies as a placeholder rather than a blank — so
    # without this the fallback "rescues" them into a row of zeroes.
    if (model.lower() in _PLACEHOLDERS
            and clean(record.get("S.N")).lower() in _PLACEHOLDERS):
        return False
    return True


#: The labels these forms print beside each field.
_FIELD_LABEL_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("S.N", re.compile(r"^serial\s*(no\.?|number)?\s*:?\s*$|^s\.?\s*n\.?\s*:?\s*$", re.I)),
    ("Model", re.compile(r"^model\s*:?\s*$", re.I)),
    ("Manufacturer", re.compile(r"^manufacturer\s*:?\s*$", re.I)),
    ("Location", re.compile(r"^location\s*:?\s*$", re.I)),
]

#: Headings that introduce the instrument doing the calibrating.
#:
#: This is not a nicety. A Hemodialysis cover page prints the reference meter
#: ABOVE the device under test:
#:
#:     A15  Calibration Device   G15 Model   H15 EMIS      <- the calibrator
#:     A22  Device information
#:     A24  Device name          G24 Model   H24 AK96      <- the actual device
#:
#: so "take the first Model: label" records the calibrator's model and serial
#: as the dialysis machine's. Position cannot be trusted; the headings can.
_GEAR_HEADING = re.compile(
    r"calibration\s*device|reference\s*(meter|equipment|instrument|standard)"
    r"|test\s*equipment|standard\s*used|m\s*&\s*te|calibrator\b|equipment\s*used"
    r"|measuring\s*equipment|traceab",
    re.I,
)
#: The heading that opens the block describing the device being calibrated.
_DEVICE_HEADING = re.compile(r"device\s*(information|data|under\s*test)|equipment\s*under", re.I)

_LABEL_COLUMNS = "ABCDEFGHIJKLMN"
_LABEL_MAX_ROW = 95

#: Other things printed on these sheets that are captions, never values.
_OTHER_CAPTIONS = re.compile(
    r"^\s*(date\s*(of\s*receipt)?|issue\s*date.*|status|type|class|"
    r"prev\.?\s*calib\.?|next\s*calib\.?.*|tested\s*by|entered\s*by|"
    r"revised\s*by|safety|remark s?|accessories)\s*:?\s*$",
    re.I,
)


def _is_a_label(text: str) -> bool:
    """True when a cell holds a caption rather than an answer."""
    return (any(pattern.match(text) for _, pattern in _FIELD_LABEL_PATTERNS)
            or bool(_OTHER_CAPTIONS.match(text)))


def locate_by_labels(values: dict[str, object]) -> dict[str, str]:
    """Work out where the fields sit from the form's own printed labels.

    ``values`` is a whole-grid read of A1:N95. Returns a cell map, which may be
    partial — a caller should check it produced a plausible record before
    trusting it.

    Only cells inside the device's own section are considered; see
    ``_GEAR_HEADING``. Where the form has no such headings at all, the first
    match top-down wins, which is what every single-block form wants.
    """
    text = {ref: clean(value) for ref, value in values.items()}

    device_row = gear_row = None
    for row in range(1, _LABEL_MAX_ROW + 1):
        for column in _LABEL_COLUMNS:
            body = text.get(f"{column}{row}", "")
            if not body:
                continue
            if device_row is None and _DEVICE_HEADING.search(body):
                device_row = row
            if _GEAR_HEADING.search(body) and (gear_row is None or row < gear_row):
                gear_row = row

    def in_device_section(row: int) -> bool:
        if device_row is not None:
            # The device block runs from its heading to the next gear heading
            # below it, or to the end of the sheet.
            end = gear_row if gear_row is not None and gear_row > device_row else None
            return row >= device_row and (end is None or row < end)
        if gear_row is not None:
            return row < gear_row
        return True

    found: dict[str, str] = {}
    for row in range(1, _LABEL_MAX_ROW + 1):
        if not in_device_section(row):
            continue
        for index, column in enumerate(_LABEL_COLUMNS):
            body = text.get(f"{column}{row}", "")
            if not body:
                continue
            for field, pattern in _FIELD_LABEL_PATTERNS:
                if field in found or not pattern.match(body):
                    continue
                # Merges make a label repeat across its own span, so the value
                # is the first cell to the right whose text differs — and which
                # is not itself a label. The Therapeutic Ultrasound form heads a
                # table with "Model | S.N." at row 11 and puts the real fields
                # at row 69; without this the model reads as the string "S.N.".
                for other in _LABEL_COLUMNS[index + 1:]:
                    candidate = text.get(f"{other}{row}", "")
                    if not candidate or candidate == body:
                        continue
                    if _is_a_label(candidate):
                        break
                    found[field] = f"{other}{row}"
                    break
    return found


def _uniform_offset(configured: dict[str, str],
                    located: dict[str, str]) -> int | None:
    """How far the form shifted, when every located field moved together.

    An inserted row moves *everything* below it, so a form whose Model, serial,
    manufacturer and location all sit one row lower has moved its Date and
    Status too. Those two carry no label the locator can match — Status is
    captioned above its value rather than beside it — so they can only be
    reached by applying the offset the other four agree on.

    Returns None unless the agreement is exact: same column, same distance, for
    every field found. A partial or ragged match means something other than a
    shift, and guessing there is how a caption ends up in the register.
    """
    offsets = set()
    for field, ref in located.items():
        old = configured.get(field)
        if not old:
            return None
        old_column, old_row = coordinate_to_tuple(old)[1], coordinate_to_tuple(old)[0]
        new_column, new_row = coordinate_to_tuple(ref)[1], coordinate_to_tuple(ref)[0]
        if old_column != new_column:
            return None
        offsets.add(new_row - old_row)
    if len(offsets) == 1:
        moved = offsets.pop()
        return moved or None
    return None


def _shift(ref: str, rows: int) -> str:
    row, column = coordinate_to_tuple(ref)
    return f"{get_column_letter(column)}{row + rows}"


def _read_with(source, cell_map: dict[str, str]) -> Record:
    values = source.values({ref for ref in cell_map.values() if ref})
    return {field: clean(values.get(ref)) if ref else ""
            for field, ref in cell_map.items()}


def read_best(filepath: str, config: dict,
              extra: dict[str, str] | None = None) -> tuple[Record, str]:
    """Read a form, falling back through every layout we know before giving up.

    Returns the record and how it was obtained: ``"primary"``, ``"alt N"``,
    ``"labels"``, ``"labels on <sheet>"``, or ``"none"``.

    **The configured map always wins when it produces a plausible record**, so
    this can only rescue a file the map got wrong; it can never change one it
    already got right. Every later step costs an extra read, and 94.7% of forms
    stop at the first.

    ``extra`` is merged into every read, for a caller that wants more of the
    sheet in the same pass — firebase_export takes the client header that way
    rather than opening all 47,000 workbooks twice.
    """
    cells = {**config["cells"], **(extra or {})}
    source = _open_workbook(filepath)
    try:
        record = _read_with(source, cells)
        if plausible(record):
            return record, "primary"

        for index, alternate in enumerate(config.get("alt_cells", []), start=1):
            candidate = _read_with(source, {**alternate, **(extra or {})})
            if plausible(candidate):
                return candidate, f"alt {index}"

        # Nothing written down fits. Ask the form where its fields are.
        grid = {f"{c}{r}": f"{c}{r}"
                for r in range(1, _LABEL_MAX_ROW + 1) for c in _LABEL_COLUMNS}
        located = locate_by_labels(source.values(set(grid)))
        if located:
            candidate = {**cells, **located}
            # Carry Date and Status along when the form has simply shifted:
            # they have no label to find, and leaving them behind is how the
            # caption "Status:" ends up in the register instead of "Calibrated".
            rows = _uniform_offset(cells, located)
            if rows is not None:
                for field in ("Date", "Status", "Status2", "S.N2"):
                    if cells.get(field):
                        candidate[field] = _shift(cells[field], rows)
            found = _read_with(source, candidate)
            if plausible(found):
                return found, "labels"

        # Still nothing: the device details may be on another tab entirely.
        for name in getattr(source, "sheet_names", []):
            if name == getattr(source, "sheet_name", None):
                continue
            if not source.select_sheet(name):
                continue
            located = locate_by_labels(source.values(set(grid)))
            if not located:
                continue
            # Only what was actually located on THIS sheet. The configured
            # references describe a different tab, so carrying them across
            # reads whatever happens to sit at those coordinates here — which
            # is how a Nebulizer's date came back as "Gas Flow Analyser".
            # A blank field is honest; a confident wrong one is not.
            found = _read_with(source, {field: located.get(field, "")
                                        for field in cells})
            if plausible(found):
                return found, f"labels on {name!r}"

        return record, "none"
    finally:
        source.close()


# ──────────────────────────────────────────────────────────────────────────────
# Building records
# ──────────────────────────────────────────────────────────────────────────────

def build_second_row(record: Record, second_row: dict) -> Record:
    """Derive the sub-module row that accompanies a two-row device.

    Same data as the parent, but with its own device name, its status taken
    from the parent's Status2 cell, and its code token rewritten.
    """
    extra = dict(record)
    extra["Device"] = second_row["device_name"]
    extra["Status"] = record.get("Status2", "")
    extra["_row_order"] = 1

    old_token, new_token = second_row["code_replace"]
    # count=1: replace only the device code, never a site prefix that happens
    # to contain the same letters.
    extra["Code"] = re.sub(re.escape(old_token), new_token, extra["Code"], count=1)
    return extra


#: In quiet mode, how many files pass between progress lines. Low enough that
#: a stalled run still looks different from a fast one, high enough that forty
#: thousand forms produce a readable log rather than forty thousand lines.
HEARTBEAT_EVERY = 500


def _heartbeat(index: int, total: int, started: float) -> None:
    elapsed = time.monotonic() - started
    rate = index / elapsed if elapsed > 0 else 0
    log.info("        read %s of %s  ·  %.0f/s", f"{index:,}", f"{total:,}", rate)


def extract_records(
    source_files: Iterable[str],
    on_file: ProgressHook | None = None,
    cancel: threading.Event | None = None,
    strict_names: bool = False,
    quiet: bool = False,
) -> tuple[list[Record], list[FileOutcome]]:
    """Read every source file, returning the records and a per-file outcome.

    ``on_file(outcome, index, total)`` fires after each file. ``cancel`` is
    checked between files, so a long run can be stopped without waiting for it
    to finish; files not reached are reported as CANCELLED. ``strict_names``
    enforces the house filename format, skipping anything that breaks it.

    ``quiet`` drops the line-per-file commentary in favour of a heartbeat.
    Problems are still reported in full — it is the forty thousand successes
    that make a log unreadable, never the failures.
    """
    ordered = sorted(source_files)
    total = len(ordered)
    records: list[Record] = []
    outcomes: list[FileOutcome] = []
    started = time.monotonic()

    def report(outcome: FileOutcome, index: int) -> None:
        outcomes.append(outcome)
        if on_file:
            on_file(outcome, index, total)

    for index, filepath in enumerate(ordered, start=1):
        if cancel is not None and cancel.is_set():
            for remaining in ordered[index - 1:]:
                outcomes.append(FileOutcome(os.path.basename(remaining), remaining,
                                            CANCELLED, detail="Run cancelled"))
            log.warning("Cancelled — %d file(s) not processed.", total - index + 1)
            break

        filename = os.path.basename(filepath)
        pre = classify_file(filepath, strict_names)

        if pre.status == BAD_FORMAT:
            log.error("%s — %s; file skipped", filename, pre.detail)
            report(pre, index)
            continue

        if pre.status == UNSUPPORTED:
            log.warning("%s — %s", filename, pre.detail)
            report(pre, index)
            continue

        if pre.status == UNKNOWN_CODE:
            if SKIP_UNKNOWN_CODES:
                log.error("%s — code '%s' is not in device_config.py; file skipped",
                          filename, pre.device_code)
                report(pre, index)
                continue
            log.warning("%s — code '%s' not found, using fallback cell map",
                        filename, pre.device_code)
            config = UNKNOWN_FALLBACK
        else:
            config = DEVICE_CONFIGS[pre.device_code]

        try:
            record, layout = read_best(filepath, config)
            if layout not in ("primary", "none"):
                # Worth saying out loud: a form reaching here has been
                # re-laid-out since its map was written, and the register only
                # has a row for it because the fallback found the fields.
                log.info("%s — read using %s (the form has moved since its "
                         "cell map was written)", filename, layout)
            record["Code"] = Path(filename).stem
            record["Device"] = config["device_name"]

            # Devices with a probe or tube carry a second serial; show both.
            if record.get("S.N2", "").strip():
                record["S.N"] = f"{record['S.N']}\n({record['S.N2']})"

            # Sort keys, so ordering never has to re-derive the code tokens.
            record["_group"] = record["Code"]
            record["_row_order"] = 0
            # Kept apart from Code because build_second_row rewrites that.
            record["_source"] = filename

            records.append(record)
            rows = 1
            if quiet:
                if index % HEARTBEAT_EVERY == 0:
                    _heartbeat(index, total, started)
            else:
                log.info(
                    "[OK]    %s  →  %s (%s)  |  Model: %s  |  S/N: %s  |  Status: %s",
                    filename, config["device_name"], pre.device_code,
                    record.get("Model", ""), record.get("S.N", ""),
                    record.get("Status", ""),
                )

            if "second_row" in config:
                extra = build_second_row(record, config["second_row"])
                records.append(extra)
                rows = 2
                if not quiet:
                    log.info(
                        "        ↳ 2nd row added: %s  |  Code: %s  |  Status: %s",
                        extra["Device"], extra["Code"], extra["Status"],
                    )

            # A dual-serial device carries its second number on its own line
            # for the spreadsheet; a table row wants it on one.
            serial = record.get("S.N", "").replace("\n", "  /  ")
            report(FileOutcome(filename, filepath, OK, pre.device_code,
                               config["device_name"], serial=serial,
                               rows=rows), index)

        except Exception as exc:
            log.error("%s — %s", filename, exc)
            report(FileOutcome(filename, filepath, ERROR, pre.device_code,
                               pre.device_name, detail=str(exc)), index)

    return records, outcomes


# ──────────────────────────────────────────────────────────────────────────────
# Ordering and de-duplication
# ──────────────────────────────────────────────────────────────────────────────

def natural_key(text: str) -> tuple:
    """Sort key that orders embedded numbers numerically (AGH9 before AGH10).

    Each part is tagged with a type rank so text and numbers never end up being
    compared against each other.
    """
    return tuple(
        (1, int(part)) if part.isdigit() else (0, part)
        for part in re.split(r"(\d+)", text.lower())
    )


def sort_records(records: list[Record]) -> list[Record]:
    """Order by parent code, keeping each generated sub-module row behind it."""
    return sorted(records, key=lambda r: (natural_key(r.get("_group", "")),
                                          r.get("_row_order", 0)))


def source_name(record: Record) -> str:
    """The file a record came from, for messages that send the user to it.

    ``_source`` is set when the record is read and survives into a generated
    sub-module row, which is the case ``Code`` cannot cover: build_second_row
    rewrites the code token, so a sub-module's Code names no file on disk.
    """
    return record.get("_source") or record.get("Code", "?")


def deduplicate_records(records: list[Record],
                        dropped: list[Duplicate] | None = None) -> list[Record]:
    """Drop records that repeat a serial number.

    A device and its generated sub-module row are allowed to share one serial
    (they are the same physical unit); any third record with that serial is
    still dropped. Blank serials are always kept.

    What gets logged is the two **filenames** involved, because that is what
    the user has to go and open. The device type does not identify which form
    to look at when a round holds a dozen of the same model.
    """
    seen: dict[str, list[tuple[str, str]]] = {}      # serial → [(device, file)]
    kept: list[Record] = []

    for record in records:
        serial = record.get("S.N", "").strip()
        device = record.get("Device", "")
        source = source_name(record)

        if not serial:
            kept.append(record)
            continue

        if serial not in seen:
            seen[serial] = [(device, source)]
            kept.append(record)
            continue

        existing = seen[serial]
        devices = [d for d, _ in existing]
        if (len(existing) == 1
                and frozenset(devices + [device]) in ALLOWED_SHARED_SN_PAIRS):
            existing.append((device, source))
            kept.append(record)
        else:
            first = existing[0][1]
            same = " (the same file)" if first == source else ""
            log.warning(
                "Duplicate serial '%s' — skipped %s, already recorded by %s%s",
                serial, source, first, same)
            if dropped is not None:
                dropped.append(Duplicate(serial, device, source, first))

    return kept


# ──────────────────────────────────────────────────────────────────────────────
# Writing output
# ──────────────────────────────────────────────────────────────────────────────

def resolve_output_path(source_files: list[str], template_file: str,
                        output_dir: str | os.PathLike | None = None) -> Path:
    """Pick where the output goes, refusing to write over the template.

    ``output_dir`` is the folder the user chose; without one the register lands
    beside the first source file, which is the long-standing default.

    Windows paths are case-insensitive, so an output called "device list.xlsx"
    would silently overwrite a template named "Device List.xlsx" sitting in the
    same folder.
    """
    folder = Path(output_dir) if output_dir else Path(source_files[0]).parent
    output_path = folder / OUTPUT_NAME
    same_file = (os.path.normcase(os.path.abspath(output_path))
                 == os.path.normcase(os.path.abspath(template_file)))
    if same_file:
        raise ValueError(
            f"The output file ({OUTPUT_NAME}) would overwrite the template. "
            f"Move the template out of {output_path.parent}, or rename it."
        )
    return output_path


def write_output(records: list[Record], template_file: str, output_path: Path) -> None:
    """Fill a copy of the template with the records and save it.

    On a Turbo-scale run this is the long tail: ~2.5s to fill 38,000 rows and
    ~7s for openpyxl to serialise them. A GUI caller sees stalls of up to half
    a second in there — measured, and they sit inside workbook.save(), so
    yielding the GIL around the fill loop does not touch them (tried; it moved
    the worst stall 483ms → 416ms and cost a second on the write). What the
    window needs is therefore to *say* what it is doing, which process_files
    logs before calling this, not to pretend the pause is not happening.
    """
    workbook = load_workbook(template_file)
    try:
        sheet = workbook.active
        for index, record in enumerate(records):
            row = TEMPLATE_START_ROW + index
            sheet.cell(row=row, column=1, value=index + 1)
            for offset, name in enumerate(FIELDS):
                sheet.cell(row=row, column=TEMPLATE_START_COL + offset,
                           value=record.get(name, ""))

        stamp_attribution(sheet, workbook, len(records))
        workbook.save(output_path)
    finally:
        workbook.close()


def stamp_attribution(sheet, workbook, record_count: int) -> None:
    """Sign the register, visibly and in the file's own properties.

    A register gets emailed, printed and filed long after the app that made it
    is out of sight, so the credit belongs in the document rather than only in
    the tool. One line, a blank row clear of the data, plus the Excel document
    properties that show under File → Properties.
    """
    row = TEMPLATE_START_ROW + record_count + 1
    made_on = datetime.now().strftime("%d/%m/%Y")

    cell = sheet.cell(row=row, column=TEMPLATE_START_COL)
    cell.value = f"Generated by Calist — {ATTRIBUTION} — {made_on}"
    cell.font = Font(italic=True, size=9, color="FF808080")

    workbook.properties.creator = ATTRIBUTION
    workbook.properties.lastModifiedBy = ATTRIBUTION
    workbook.properties.title = "Equipment register"
    workbook.properties.description = (
        f"Compiled from {record_count} row(s) by Calist — {ATTRIBUTION}")


# ──────────────────────────────────────────────────────────────────────────────
# Orchestration
# ──────────────────────────────────────────────────────────────────────────────

def process_files(
    source_files: list[str],
    template_file: str,
    *,
    deduplicate: bool = False,
    strict_names: bool = False,
    output_dir: str | os.PathLike | None = None,
    on_file: ProgressHook | None = None,
    cancel: threading.Event | None = None,
    quiet: bool = False,
) -> RunResult:
    """Run the full pipeline and report what happened.

    Always returns a RunResult; check ``.succeeded`` or ``.output_path``. A
    cancelled run writes nothing and comes back with ``cancelled=True``.
    ``strict_names`` skips any file whose name breaks the house format.
    ``output_dir`` overrides where the register is written. ``quiet`` trades
    the line-per-file log for a heartbeat, which is what makes a run of tens of
    thousands readable.
    """
    rule = "─" * 55
    result = RunResult()

    if not source_files:
        result.error = "No source files selected."
        log.error("%s", result.error)
        return result

    log.info(rule)
    log.info("Template : %s", os.path.basename(template_file))
    log.info("Sources  : %d file(s) selected", len(source_files))
    log.info(rule)

    try:
        output_path = resolve_output_path(source_files, template_file, output_dir)
    except ValueError as exc:
        result.error = str(exc)
        log.error("%s", exc)
        return result

    records, result.outcomes = extract_records(source_files, on_file, cancel,
                                               strict_names, quiet)
    result.second_rows_added = sum(1 for o in result.outcomes if o.rows == 2)
    records = sort_records(records)

    if cancel is not None and cancel.is_set():
        result.cancelled = True
        log.warning("Run cancelled — no file written.")
        return result

    if deduplicate:
        before = len(records)
        records = deduplicate_records(records, result.duplicates)
        result.duplicates_removed = before - len(records)
        log.info("%d duplicate record(s) removed.", result.duplicates_removed) \
            if result.duplicates_removed else log.info("No duplicates found.")

    if not records:
        result.error = "No valid records extracted. No file created."
        log.warning("%s", result.error)
        return result

    log.info("Writing %s row(s)…", f"{len(records):,}")
    try:
        write_output(records, template_file, output_path)
    except Exception as exc:
        result.error = f"Could not save output file: {exc}"
        log.error("%s", result.error)
        return result

    result.output_path = output_path
    result.rows_written = len(records)

    log.info(rule)
    log.info("✔ Success! %d record(s) written.", len(records))
    log.info("NEW FILE SAVED AT: %s", output_path)
    return result


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def inspect_form(filepath: str) -> int:
    """Print what every mapped cell of one form actually reads.

    The answer to "why is this field blank?". Cell maps are written by reading
    coordinates off a paper form, and a form that has been re-laid-out since
    shows up here immediately: a field reading '' next to a cell reference is
    a map that no longer matches the file.

    Returns a process exit code, so a form with unreadable fields fails loudly
    when this is run from a script.
    """
    path = Path(filepath)
    outcome = classify_file(filepath)
    print(f"file   : {path.name}")
    print(f"device : {outcome.device_name or '—'} ({outcome.device_code or '—'})")

    if outcome.status != READY:
        print(f"\nCannot inspect: {outcome.detail}")
        return 1

    config = DEVICE_CONFIGS[outcome.device_code]
    cells = config["cells"]
    record = read_record(filepath, cells)
    blank = [f for f, v in record.items() if not v]

    print(f"\n{'field':<14}{'cell':<7}value")
    print("─" * 60)
    for field, ref in cells.items():
        value = record[field].replace("\n", " / ")
        print(f"{field:<14}{ref:<7}{value if value else '(blank)'}")

    # What the pipeline would actually record, which is not the same thing once
    # the configured map stops fitting the form.
    best, layout = read_best(filepath, config)
    if layout != "primary":
        print(f"\nthe configured map does not fit this form — read using: {layout}")
        if layout == "none":
            print("no layout, alternate or printed label produced a usable record")
        else:
            for field in ("Model", "S.N", "Manufacturer", "Location"):
                if best.get(field):
                    print(f"  {field:<14}{best[field]}")

    if path.suffix.lower() != ".xls":
        # Asked of the reader itself, not of a second opinion: the whole point
        # of this dump is to say what the pipeline saw.
        source = _open_workbook(filepath)
        try:
            merges = source.merged_refs()
            names = source.sheet_names
            print(f"\nsheet read: {source.sheet_name!r} "
                  f"(of {len(names)}: {', '.join(names[:6])}"
                  f"{' …' if len(names) > 6 else ''})")
            if source.skipped_sheets:
                print("skipped empty leading tab(s): "
                      f"{', '.join(source.skipped_sheets[:3])}")
            print(f"merged ranges: {len(merges)}")
            if merges:
                print("  " + ", ".join(merges[:24])
                      + (" …" if len(merges) > 24 else ""))
        finally:
            source.close()

    if blank:
        print(f"\n{len(blank)} field(s) read blank: {', '.join(blank)}")
        print("Either the form leaves them empty, or device_config.py points at "
              "the wrong cell for this layout.")
        return 1

    print("\nEvery mapped field read a value.")
    return 0


def main() -> None:
    """Launch the desktop app, or inspect a single form.

        python calist.py                    launch the app
        python calist.py --inspect FORM     dump what each mapped cell reads

    The UI is imported inside the launch branch rather than at module scope so
    that importing this module — from a test, a script, or another tool —
    costs nothing and pulls in no GUI toolkit. --inspect keeps that property.
    """
    args = sys.argv[1:]
    if args and args[0] == "--inspect":
        if len(args) != 2:
            print("usage: python calist.py --inspect <form.xlsx>")
            raise SystemExit(2)
        logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")
        raise SystemExit(inspect_form(args[1]))

    from ui import run
    run()


if __name__ == "__main__":
    main()
